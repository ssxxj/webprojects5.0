import contextlib
import io
import json
import shutil
import sys
import tempfile
import types
import unittest
from pathlib import Path
from unittest.mock import patch

import yaml


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RUNTIME_ROOT = PROJECT_ROOT / "30_runtime"
sys.path.insert(0, str(RUNTIME_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "40_evaluation" / "runtime"))
sys.path.insert(0, str(PROJECT_ROOT / "50_assets" / "课堂实施方案" / "runtime"))

import build_all_chapter_assets_v5 as build_all  # noqa: E402
import build_chapter_assets_v5 as build_single  # noqa: E402
import check_all_chapter_asset_drift_v5 as drift_check  # noqa: E402
import generate_lesson_script_v5 as lesson_gen  # noqa: E402
import preflight_projects5_v5 as preflight  # noqa: E402
import course_assignment_eval_v5 as eval_engine  # noqa: E402


FIXTURE = PROJECT_ROOT / "tests" / "fixtures" / "chapter99_demo.yaml"


class Projects5RuntimeTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.temp_root = Path(self.temp_dir.name) / "projects5.0"
        (self.temp_root / "30_runtime" / "chapter_master_configs").mkdir(parents=True, exist_ok=True)
        (self.temp_root / "40_evaluation" / "runtime" / "chapter_profiles").mkdir(parents=True, exist_ok=True)
        (self.temp_root / "50_assets" / "assignment_packs").mkdir(parents=True, exist_ok=True)
        (self.temp_root / "50_assets" / "课堂实施方案").mkdir(parents=True, exist_ok=True)
        (self.temp_root / "50_assets" / "章节讲义").mkdir(parents=True, exist_ok=True)
        self.config_path = self.temp_root / "30_runtime" / "chapter_master_configs" / "chapter99_demo.yaml"
        shutil.copyfile(FIXTURE, self.config_path)
        self.config_data = yaml.safe_load(self.config_path.read_text(encoding="utf-8"))
        self.chapter_id = self.config_data["chapter_id"]
        self.chapter_title = self.config_data["chapter_title"]
        self._write_minimal_lecture_assets()

        self.patchers = [
            patch.object(build_single, "PROJECT_ROOT", self.temp_root),
            patch.object(build_all, "PROJECT_ROOT", self.temp_root),
            patch.object(drift_check, "PROJECT_ROOT", self.temp_root),
            patch.object(lesson_gen, "PROJECT_ROOT", self.temp_root),
            patch.object(preflight, "PROJECT_ROOT", self.temp_root),
            patch.object(preflight, "LECTURE_ROOT", self.temp_root / "50_assets" / "章节讲义"),
        ]
        for patcher in self.patchers:
            patcher.start()

    def tearDown(self) -> None:
        for patcher in reversed(self.patchers):
            patcher.stop()
        self.temp_dir.cleanup()

    def _run_main_quiet(self, module, argv: list[str]) -> int:
        stdout = io.StringIO()
        stderr = io.StringIO()
        with (
            patch.object(sys, "argv", argv),
            contextlib.redirect_stdout(stdout),
            contextlib.redirect_stderr(stderr),
        ):
            return module.main()

    def _run_build_single(self) -> int:
        return self._run_main_quiet(
            build_single,
            ["build_chapter_assets_v5.py", "--input", str(self.config_path)],
        )

    def _write_minimal_lecture_assets(self) -> None:
        lecture_dir = self.temp_root / "50_assets" / "章节讲义" / self.chapter_id
        lecture_dir.mkdir(parents=True, exist_ok=True)

        teacher_path = lecture_dir / f"{self.chapter_title}_教师版讲义_v5.0.md"
        student_path = lecture_dir / f"{self.chapter_title}_学生预习版讲义_v5.0.md"
        checklist_path = lecture_dir / f"{self.chapter_title}_讲义迁移清单_v5.0.json"
        mapping_path = lecture_dir / f"{self.chapter_title}_4.0讲义到5.0讲义到5.0作业包对照表.md"

        teacher_path.write_text(f"# {self.chapter_title} 教师版讲义 v5.0\n", encoding="utf-8")
        student_path.write_text(f"# {self.chapter_title} 学生预习版讲义 v5.0\n", encoding="utf-8")
        mapping_path.write_text(f"# {self.chapter_title} 4.0讲义到5.0讲义到5.0作业包对照表\n", encoding="utf-8")
        checklist_path.write_text(
            json.dumps(
                {
                    "chapter_id": self.chapter_id,
                    "chapter_name": self.chapter_title,
                    "upstream_5_0": {
                        "profile": str(self.temp_root / "40_evaluation" / "runtime" / "chapter_profiles" / f"{self.chapter_id}.json"),
                        "assignment_pack": str(
                            self.temp_root
                            / "50_assets"
                            / "assignment_packs"
                            / self.chapter_id
                            / f"{self.chapter_title}_教师端作业包_v5.0.md"
                        ),
                    },
                    "generated_assets": [
                        str(teacher_path),
                        str(student_path),
                        str(mapping_path),
                    ],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

    def test_build_single_generates_expected_assets(self) -> None:
        code = self._run_build_single()
        self.assertEqual(code, 0)

        profile_path = self.temp_root / "40_evaluation" / "runtime" / "chapter_profiles" / "chapter99_demo.json"
        teacher_pack = self.temp_root / "50_assets" / "assignment_packs" / "chapter99_demo" / "第九十九章 Demo章节_教师端作业包_v5.0.md"
        lesson_input = self.temp_root / "50_assets" / "课堂实施方案" / "chapter99_demo" / "lesson_script_input_v5.yaml"
        lesson_script = self.temp_root / "50_assets" / "课堂实施方案" / "chapter99_demo" / "第九十九章 Demo章节_课堂实施方案_v5.0.md"

        self.assertTrue(profile_path.exists())
        self.assertTrue(teacher_pack.exists())
        self.assertTrue(lesson_input.exists())
        self.assertTrue(lesson_script.exists())
        self.assertIn("教师端作业包", teacher_pack.read_text(encoding="utf-8"))

    def test_generated_manifests_use_project_relative_paths(self) -> None:
        self.assertEqual(self._run_build_single(), 0)
        manifest_path = self.config_path.with_name("chapter99_demo_build_manifest_v5.json")
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest["input_file"], "30_runtime/chapter_master_configs/chapter99_demo.yaml")
        self.assertEqual(manifest["profile_file"], "40_evaluation/runtime/chapter_profiles/chapter99_demo.json")
        self.assertFalse(str(self.temp_root) in json.dumps(manifest, ensure_ascii=False))

        assignment_manifest = (
            self.temp_root
            / "50_assets"
            / "assignment_packs"
            / "chapter99_demo"
            / "第九十九章 Demo章节_生成清单_v5.0.json"
        )
        assignment_data = json.loads(assignment_manifest.read_text(encoding="utf-8"))
        self.assertTrue(assignment_data["generated_files"]["teacher_pack"].startswith("50_assets/assignment_packs/"))
        self.assertFalse(str(self.temp_root) in json.dumps(assignment_data, ensure_ascii=False))

    def test_build_all_validate_only_passes(self) -> None:
        code = self._run_main_quiet(
            build_all,
            ["build_all_chapter_assets_v5.py", "--root", str(self.config_path.parent), "--validate-only"],
        )
        self.assertEqual(code, 0)
        summary_path = self.config_path.parent / "all_chapters_build_manifest_v5.json"
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["success"], 1)

    def test_drift_check_detects_modified_teacher_pack(self) -> None:
        self.assertEqual(self._run_build_single(), 0)
        teacher_pack = self.temp_root / "50_assets" / "assignment_packs" / "chapter99_demo" / "第九十九章 Demo章节_教师端作业包_v5.0.md"
        teacher_pack.write_text("BROKEN\n", encoding="utf-8")

        result = drift_check.process_config(self.config_path, diff_lines=4)
        self.assertEqual(result["status"], "drift")
        self.assertIn("teacher_pack", result["drifted_files"])

    def test_schema_version_rejects_unknown_master_config(self) -> None:
        self.config_data["schema_version"] = "v6.0"
        self.config_path.write_text(yaml.safe_dump(self.config_data, allow_unicode=True, sort_keys=False), encoding="utf-8")
        errors = build_single.validate_master_config(build_single.load_master_config(self.config_path))
        self.assertTrue(any("不支持的 schema_version" in item for item in errors))

    def test_sensitive_scan_is_review_only(self) -> None:
        self.assertEqual(self._run_build_single(), 0)
        profile = eval_engine.load_profile(
            self.temp_root / "40_evaluation" / "runtime" / "chapter_profiles" / "chapter99_demo.json"
        )
        text = (
            "任务1 Demo：已经完成。授权边界：仅限 DVWA 和 localhost。"
            "学生记录了疑似真实邮箱 test_user@realcorp.com 和手机号 13800138000，"
            "但这里不应自动封顶，只应进入教师复核提醒。"
        )
        base_score, _base_debug = eval_engine.generic_fallback_scoring(text, 3, profile, None)
        score, debug = eval_engine.score_text(text, 3, profile)
        self.assertEqual(score, base_score)
        self.assertTrue(debug["sensitive_scan"]["has_findings"])
        self.assertEqual(debug["sensitive_scan"]["mode"], "review_only")
        self.assertNotIn("敏感信息疑似未脱敏", debug.get("hard_gate_reasons", []))

    def test_excel_update_writes_scores_and_feedback(self) -> None:
        from openpyxl import Workbook, load_workbook

        profile = eval_engine.ChapterProfile(
            course_name="Web应用安全与防护",
            chapter_name="测试章",
            chapter_mainline="测试主线",
            capability_goals=["目标"],
            tasks=[eval_engine.TaskRule(name="任务", score=10, required=True, semantic_requirements=["字段"])],
            relation_item_name="关系图",
            relation_item_score=5,
            self_eval_score=5,
            redlines=[eval_engine.RedlineRule(name="未脱敏", description="敏感信息未脱敏", action="lt_60")],
            professional_checks=["边界"],
            default_tags=["T1"],
        )
        wb = Workbook()
        ws = wb.active
        ws.append(["学号", "姓名"])
        ws.append(["01", "张三"])
        excel_path = Path(self.temp_dir.name) / "roster.xlsx"
        wb.save(excel_path)

        outcome = eval_engine.SubmissionOutcome(
            file_path=None,
            student_id="01",
            student_name="张三",
            page_count=2,
            submitted=True,
            score=88,
            label="良好",
            assessment="88分，良好。",
            good_points="证据清楚。",
            suggestion="补强机制解释。",
            debug={},
        )
        eval_engine.update_excel(excel_path, [outcome], profile)
        updated = load_workbook(excel_path).active
        self.assertEqual(updated.cell(row=1, column=3).value, "测试章 v5.0分数")
        self.assertEqual(updated.cell(row=2, column=3).value, 88)
        self.assertEqual(updated.cell(row=2, column=6).value, "补强机制解释。")

    def test_optional_rapidocr_import_failure_is_non_fatal(self) -> None:
        with patch.object(eval_engine.importlib, "import_module", side_effect=ImportError("missing rapidocr")):
            extractor = eval_engine.OCRExtractor()
        self.assertFalse(extractor.ocr_available)
        self.assertIsNone(extractor.ocr)
        self.assertIn("RapidOCR 初始化失败", extractor.ocr_error)

    def test_ocr_runtime_error_falls_back_to_page_text(self) -> None:
        class FakePix:
            def tobytes(self, _fmt: str) -> bytes:
                return b"png"

        class FakePage:
            def get_text(self, _mode: str) -> str:
                return "短文本"

            def get_pixmap(self, matrix=None, alpha=False):  # noqa: ARG002
                return FakePix()

        class FakeDoc:
            page_count = 1

            def __iter__(self):
                return iter([FakePage()])

            def close(self) -> None:
                return None

        def fake_ocr(_img):
            raise RuntimeError("boom")

        fake_open = lambda _path: FakeDoc()
        fake_fitz = types.SimpleNamespace(open=fake_open, Matrix=eval_engine.fitz.Matrix)

        with patch.object(eval_engine, "fitz", fake_fitz):
            extractor = eval_engine.OCRExtractor()
            extractor.ocr = fake_ocr
            extractor.ocr_available = True
            text, page_count = extractor.extract_pdf_text(Path("demo.pdf"))

        self.assertEqual(page_count, 1)
        self.assertEqual(text, "短文本")
        self.assertFalse(extractor.ocr_available)
        self.assertIn("OCR 执行失败", extractor.ocr_runtime_error)

    def test_ocr_status_dict_contains_expected_fields(self) -> None:
        extractor = eval_engine.OCRExtractor()
        status = extractor.status_dict()
        self.assertIn("available", status)
        self.assertIn("init_error", status)
        self.assertIn("runtime_error", status)

    def test_preflight_reports_release_ready_on_clean_project(self) -> None:
        self.assertEqual(self._run_build_single(), 0)
        root = self.config_path.parent
        code = self._run_main_quiet(
            preflight,
            ["preflight_projects5_v5.py", "--root", str(root)],
        )
        self.assertEqual(code, 0)
        report_path = root / "projects5_preflight_report_v5.json"
        report = json.loads(report_path.read_text(encoding="utf-8"))
        self.assertTrue(report["release_ready"])
        self.assertEqual(report["asset_consistency"]["drift"], 0)
        self.assertEqual(report["lecture_governance"]["issue"], 0)

    def test_preflight_fails_when_assets_drift(self) -> None:
        self.assertEqual(self._run_build_single(), 0)
        teacher_pack = self.temp_root / "50_assets" / "assignment_packs" / "chapter99_demo" / "第九十九章 Demo章节_教师端作业包_v5.0.md"
        teacher_pack.write_text("BROKEN\n", encoding="utf-8")
        root = self.config_path.parent
        code = self._run_main_quiet(
            preflight,
            ["preflight_projects5_v5.py", "--root", str(root)],
        )
        self.assertEqual(code, 1)
        report_path = root / "projects5_preflight_report_v5.json"
        report = json.loads(report_path.read_text(encoding="utf-8"))
        self.assertFalse(report["release_ready"])
        self.assertEqual(report["asset_consistency"]["drift"], 1)

    def test_preflight_fails_when_lecture_assets_missing(self) -> None:
        self.assertEqual(self._run_build_single(), 0)
        teacher_lecture = (
            self.temp_root
            / "50_assets"
            / "章节讲义"
            / self.chapter_id
            / f"{self.chapter_title}_教师版讲义_v5.0.md"
        )
        teacher_lecture.unlink()
        root = self.config_path.parent
        code = self._run_main_quiet(
            preflight,
            ["preflight_projects5_v5.py", "--root", str(root)],
        )
        self.assertEqual(code, 1)
        report_path = root / "projects5_preflight_report_v5.json"
        report = json.loads(report_path.read_text(encoding="utf-8"))
        self.assertFalse(report["release_ready"])
        self.assertEqual(report["lecture_governance"]["issue"], 1)


if __name__ == "__main__":
    unittest.main()
