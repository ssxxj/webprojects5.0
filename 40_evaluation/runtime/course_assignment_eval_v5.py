#!/usr/bin/env python3
"""
projects5.0 通用作业评估引擎

当前版本完成四件事：
1. 继续支持 profile 校验与摘要输出；
2. 将 projects4.0 第二章成熟评分逻辑迁入 v5 通用引擎；
3. 为第一章到第八章提供章节 profile；
4. 为第一章到第八章提供可执行 scorer，并支持单 PDF / 目录批量评分。

注意：
- 评分仍以“课程母规则 + 章节 profile + 章节 scorer”为主。
- 如后续继续新增章节，应优先复用当前 scorer 收口方式接入。
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import asdict, dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

import fitz
from openpyxl import load_workbook
from rapidocr_onnxruntime import RapidOCR
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle


URL_RE = re.compile(r"https?://[^\s)\]>\"'，。]+", re.IGNORECASE)
DOMAIN_RE = re.compile(r"\b(?:[a-z0-9-]+\.)+[a-z]{2,}\b", re.IGNORECASE)
IP_RE = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")
PRIVATE_IP_RE = re.compile(
    r"^(?:127\.0\.0\.1|localhost|10(?:\.\d{1,3}){3}|192\.168(?:\.\d{1,3}){2}|172\.(?:1[6-9]|2\d|3[01])(?:\.\d{1,3}){2})$",
    re.IGNORECASE,
)
DATE_RE = re.compile(r"\b\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}(?:日)?\b")
SESSION_LEAK_RE = re.compile(
    r"(?:jsessionid|phpsessid|sessionid|token|csrftoken|user_token|authorization)\s*[:=：]\s*[a-z0-9_\-]{8,}",
    re.IGNORECASE,
)
MASK_HINT_RE = re.compile(r"(?:\*{4,}|打码|遮挡|脱敏|隐藏)")
SCREENSHOT_MARKER_RE = r"(截图|圈出|标出|标注|图[一二三四五六七八九十]|图\d+)"
ABSOLUTE_ASSERTIONS = [
    "已经证明存在漏洞",
    "已经确认存在漏洞",
    "已经发现漏洞",
    "一定存在漏洞",
    "肯定存在漏洞",
    "直接等于漏洞",
    "开放端口就等于漏洞",
    "开放端口等于漏洞",
    "说明目标一定存在风险",
]
AI_STOCK_PHRASES = [
    "综上所述",
    "由此可见",
    "不难看出",
    "进一步说明",
    "总的来说",
    "值得注意的是",
    "在实际应用中",
    "从某种意义上说",
    "可以明显看出",
]
ABSTRACT_TERMS = [
    "信息层",
    "证据价值",
    "边界意识",
    "方法论",
    "技术证据",
    "系统性",
    "分层模型",
    "风险控制",
    "逻辑闭环",
]
SIMILARITY_REMOVALS = [
    "项目目标和授权边界",
    "授权边界说明",
    "学生自评表",
    "任务完成情况",
    "机制理解情况",
    "证据质量情况",
    "合规与边界",
    "自评结论",
    "最小机制说明",
    "ai输出审核与人工复核记录",
    "人工审核痕迹",
    "人工复核痕迹",
]
CN_NUMS = "一二三四五六七八九十"


@dataclass
class TaskRule:
    name: str
    score: float
    required: bool = True
    semantic_requirements: list[str] = field(default_factory=list)


@dataclass
class RedlineRule:
    name: str
    description: str
    action: str = "lt_60"


@dataclass
class ChapterProfile:
    course_name: str
    chapter_name: str
    chapter_mainline: str
    capability_goals: list[str]
    tasks: list[TaskRule]
    relation_item_name: str
    relation_item_score: float
    self_eval_score: float
    redlines: list[RedlineRule]
    professional_checks: list[str]
    default_tags: list[str]


@dataclass
class SubmissionOutcome:
    file_path: str | None
    student_id: str
    student_name: str
    page_count: int
    submitted: bool
    score: int
    label: str
    assessment: str
    good_points: str
    suggestion: str
    debug: dict[str, Any]


def normalize_text(text: str) -> str:
    text = text.replace("\u3000", " ")
    text = text.replace("\ufeff", "")
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def compact_text(text: str) -> str:
    return re.sub(r"\s+", "", text.lower())


def contains_any(text: str, patterns: Iterable[str]) -> bool:
    return any(pattern in text for pattern in patterns)


def count_hits(text: str, patterns: Iterable[str]) -> int:
    return sum(1 for pattern in patterns if pattern in text)


def count_regex(text: str, pattern: str) -> int:
    return len(re.findall(pattern, text, flags=re.IGNORECASE | re.MULTILINE))


def safe_snippet(text: str, limit: int = 220) -> str:
    return normalize_text(text).replace("\n", " ")[:limit]


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def ratio_score(hit_count: int, total: int, max_score: float) -> float:
    if total <= 0:
        return 0.0
    return clamp(hit_count / total, 0.0, 1.0) * max_score


def max_ratio_score(primary_checks: Iterable[bool], fallback_checks: Iterable[bool], max_score: float) -> float:
    primary_list = list(primary_checks)
    fallback_list = list(fallback_checks)
    primary_score = ratio_score(sum(primary_list), len(primary_list), max_score)
    fallback_score = ratio_score(sum(fallback_list), len(fallback_list), max_score)
    return max(primary_score, fallback_score)


def task_ratio_score(checks: Iterable[bool], max_score: float) -> float:
    checks_list = list(checks)
    return ratio_score(sum(checks_list), len(checks_list), max_score)


def count_numbered_items(text: str) -> int:
    return len(
        re.findall(
            r"(?m)^\s*(?:\d+\s*[.、:：)]|[一二三四五六七八九十]+\s*[、:：)])",
            text,
        )
    )


def extract_submission_identity(pdf_path: Path) -> tuple[str, str]:
    stem = pdf_path.stem.replace("_", " ").strip()
    stem = re.sub(r"第?[一二三四五六七八九十0-9]+章.*$", "", stem, flags=re.IGNORECASE).strip()
    match = re.match(r"^\s*(\d+)\s*(.*)$", stem)
    student_id = match.group(1).strip() if match else ""
    student_name = re.sub(r"\s+", "", match.group(2).strip() if match else stem)
    return student_id, student_name


def extract_domains(text: str) -> list[str]:
    domains: list[str] = []
    for url in URL_RE.findall(text):
        match = re.match(r"https?://([^/\s]+)", url, re.IGNORECASE)
        if match:
            domains.append(match.group(1).lower())
    for domain in DOMAIN_RE.findall(text):
        lowered = domain.lower()
        if lowered not in domains:
            domains.append(lowered)
    return domains


def is_private_target(value: str) -> bool:
    value = value.strip().lower()
    if not value:
        return False
    if PRIVATE_IP_RE.match(value):
        return True
    if value.startswith("127.0.0.1") or value.startswith("localhost"):
        return True
    if value.startswith("192.168.") or value.startswith("10."):
        return True
    if re.match(r"^172\.(?:1[6-9]|2\d|3[01])\.", value):
        return True
    return False


def find_label_value(text: str, labels: Iterable[str]) -> str:
    for label in labels:
        pattern = rf"{re.escape(label)}\s*[:：]\s*([^\n]+)"
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return ""


def compact_contains(text: str, phrases: Iterable[str]) -> int:
    compact = compact_text(text)
    return sum(1 for phrase in phrases if compact_text(phrase) in compact)


def split_paragraphs(text: str) -> list[str]:
    paragraphs: list[str] = []
    for chunk in re.split(r"\n\s*\n", text):
        clean = normalize_text(chunk)
        if len(clean) >= 40:
            paragraphs.append(clean)
    return paragraphs


def detect_ai_like_paragraphs(text: str) -> list[str]:
    paragraphs = split_paragraphs(text)
    flagged: list[str] = []
    for paragraph in paragraphs:
        stock_hits = count_hits(paragraph, AI_STOCK_PHRASES)
        abstract_hits = count_hits(paragraph, ABSTRACT_TERMS)
        concrete_hits = count_hits(
            paragraph,
            [
                "Google",
                "Whois",
                "Shodan",
                "Burp",
                "Nmap",
                "SQL",
                "DVWA",
                "XSS",
                "Cookie",
                "Session",
                "URL",
                "Method",
                "状态码",
                "端口",
                "参数化查询",
            ],
        )
        concrete_hits += 1 if URL_RE.search(paragraph) else 0
        concrete_hits += 1 if IP_RE.search(paragraph) else 0
        if len(paragraph) >= 80 and stock_hits >= 2 and abstract_hits >= 2 and concrete_hits <= 1:
            flagged.append(safe_snippet(paragraph, 110))
    return flagged[:2]


def detect_ai_review_trace(text: str) -> dict[str, Any]:
    compact = compact_text(text)
    has_section = (
        "人工审核痕迹" in text
        or "人工复核痕迹" in text
        or "AI输出审核与人工复核记录" in text
        or "ai输出审核与人工复核记录" in compact
        or "ai输出审核" in compact
    )
    has_ai_use = contains_any(compact, ["本作业是否使用ai", "未使用ai", "局部使用", "深度使用"])
    has_keep = contains_any(compact, ["保留了ai输出中的哪2处内容", "保留了ai输出中的哪两处内容", "为什么保留"])
    has_edit = contains_any(compact, ["删改了ai输出中的哪2处内容", "删改了ai输出中的哪两处内容", "为什么删改", "为什么放弃"])
    has_verified = contains_any(compact, ["自行核实过的3个专业判断", "自行核实过的三个专业判断"])
    has_boundary = contains_any(compact, ["最容易误判的一处边界问题", "最容易误判的一个边界问题"])
    item_hits = sum(1 for value in [has_ai_use, has_keep, has_edit, has_verified, has_boundary] if value)
    present = has_section or item_hits >= 2
    complete = present and item_hits >= 4
    return {
        "has_section": has_section,
        "has_ai_use": has_ai_use,
        "has_keep": has_keep,
        "has_edit": has_edit,
        "has_verified": has_verified,
        "has_boundary": has_boundary,
        "item_hits": item_hits,
        "present": present,
        "complete": complete,
    }


def clean_similarity_text(text: str) -> str:
    normalized = normalize_text(text)
    kept_lines: list[str] = []
    for raw_line in normalized.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if "：" in line or ":" in line:
            value = re.split(r"[:：]", line, maxsplit=1)[1].strip()
            if value:
                kept_lines.append(value)
        elif len(line) >= 4:
            kept_lines.append(line)
    joined = " ".join(kept_lines) if kept_lines else normalized
    for phrase in SIMILARITY_REMOVALS:
        joined = joined.replace(phrase, " ")
    joined = re.sub(r"任务\s*[1-9一二三四五六七八九十]\s*", " ", joined, flags=re.IGNORECASE)
    joined = re.sub(r"[一二三四五六七八九十]+\s*[、.．：:]\s*", " ", joined)
    joined = re.sub(r"\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}(?:日)?", " ", joined)
    joined = re.sub(r"[^\w\u4e00-\u9fff]+", "", joined.lower())
    return joined


def find_context_window(text: str, anchors: Iterable[str], window: int = 2200) -> str:
    positions: list[int] = []
    for anchor in anchors:
        match = re.search(re.escape(anchor), text, re.IGNORECASE)
        if match:
            positions.append(match.start())
    if not positions:
        return ""
    start = min(positions)
    end = min(len(text), start + window)
    return text[start:end].strip()


def page_points(page_count: int) -> float:
    if page_count >= 10:
        return 4.5
    if page_count >= 8:
        return 4.2
    if page_count >= 6:
        return 3.6
    if page_count >= 4:
        return 2.8
    if page_count >= 2:
        return 1.8
    return 0.8


def screenshot_points(screenshot_markers: int) -> float:
    if screenshot_markers >= 10:
        return 4.5
    if screenshot_markers >= 7:
        return 3.8
    if screenshot_markers >= 4:
        return 3.0
    if screenshot_markers >= 2:
        return 2.0
    if screenshot_markers >= 1:
        return 1.0
    return 0.0


def build_assessment(score: int, debug: dict[str, Any]) -> str:
    hard_gate_reasons: list[str] = debug.get("hard_gate_reasons", [])
    content_issues: list[str] = debug.get("content_issues", [])
    label = debug.get("label", label_for_score(score, hard_gate_reasons))
    if hard_gate_reasons:
        return f"{score}分，{label}；" + "；".join(hard_gate_reasons[:2]) + "。"

    strengths: list[str] = []
    if debug.get("task_completion", 0) >= 16:
        strengths.append("任务完成度较高")
    if debug.get("evidence_quality", 0) >= 13:
        strengths.append("证据较清楚")
    if debug.get("mechanism_score", 0) >= 15:
        strengths.append("机制解释较到位")
    if debug.get("professional_score", 0) >= 8:
        strengths.append("专业判断较稳妥")
    if debug.get("expression_score", 0) >= 8:
        strengths.append("结构较清楚")
    if not strengths:
        strengths.append("基本完成本章作业")
    if content_issues and score < 80:
        return f"{score}分，{label}；" + "、".join(strengths[:2]) + f"；仍需完善：{content_issues[0]}。"
    return f"{score}分，{label}；" + "、".join(strengths[:3]) + "。"


def build_good_points(debug: dict[str, Any]) -> str:
    good_points: list[str] = []
    task_scores = debug.get("task_scores", {})
    task_max = debug.get("task_max_scores", {})
    strong_tasks = [
        name for name, value in task_scores.items()
        if value >= 0.8 * task_max.get(name, 1)
    ]
    if len(strong_tasks) >= max(2, len(task_scores) - 1):
        good_points.append("主体任务完成度较高")
    if debug.get("evidence_quality", 0) >= 15:
        good_points.append("截图与关键字段定位较清楚")
    elif debug.get("evidence_quality", 0) >= 12:
        good_points.append("证据基本清楚")
    if debug.get("mechanism_score", 0) >= 16:
        good_points.append("机制解释较到位")
    if debug.get("professional_score", 0) >= 8:
        good_points.append("专业结论与边界表达较稳妥")
    if debug.get("expression_score", 0) >= 8:
        good_points.append("整体结构较清楚")
    if not good_points:
        good_points.append("已开始按章节要求组织作业内容")
    return "；".join(good_points[:3]) + "。"


def build_suggestion(profile: ChapterProfile, debug: dict[str, Any], submitted: bool) -> str:
    if not submitted:
        task_names = "、".join(task.name for task in profile.tasks)
        return f"补交完整 PDF：至少补齐 {task_names}、{profile.relation_item_name}、学生自评表和 AI输出审核与人工复核记录。"

    hard_gate_reasons: list[str] = debug.get("hard_gate_reasons", [])
    content_issues: list[str] = debug.get("content_issues", [])
    suggestions: list[str] = []
    task_scores = debug.get("task_scores", {})
    task_max_scores = debug.get("task_max_scores", {})
    task_name_map = debug.get("task_name_map", {})
    task_requirement_map = debug.get("task_requirement_map", {})

    for task_key, score in task_scores.items():
        max_score = task_max_scores.get(task_key, 0)
        if max_score and score < 0.6 * max_score:
            task_name = task_name_map.get(task_key, task_key)
            reqs = task_requirement_map.get(task_key, [])
            hint = f"重点补齐：{reqs[0]}" if reqs else "回到该任务最小字段逐项补齐"
            suggestions.append(f"补强“{task_name}”，{hint}。")

    relation_status = debug.get("relation_status", "complete")
    if relation_status == "missing":
        suggestions.append(f"补交 {profile.relation_item_name}，至少体现“关系/流程/因果/下一步”，不能只写工具或概念列表。")
    elif relation_status == "partial":
        suggestions.append(f"{profile.relation_item_name} 已提交，但关系表达仍偏弱，建议补出“输入/处理/结果/控制点/下一步”链条。")

    if not debug.get("self_eval_present", False):
        suggestions.append("补写学生自评表，说明任务完成情况、机制理解和最需要改进的一项。")
    if not debug.get("ai_review_present", False):
        suggestions.append("补写 AI输出审核与人工复核记录，明确是否使用 AI、保留与删改内容、已自行核实的专业判断。")
    if "敏感信息疑似未脱敏" in hard_gate_reasons:
        suggestions.append("重新打码 Cookie、SessionID、Token 等敏感信息后再提交。")
    if "主动探测或漏洞验证目标授权边界不清" in hard_gate_reasons:
        suggestions.append("把目标严格限制为 DVWA、127.0.0.1、localhost 或教师明确授权环境，并在文中写清授权边界。")
    if "与同班作业高度雷同" in hard_gate_reasons:
        suggestions.append("当前存在高度雷同痕迹，需提供原始完成过程并接受人工复核。")

    professional_errors = debug.get("professional_errors", [])
    if professional_errors:
        suggestions.append(f"修正专业判断：{professional_errors[0]['item']}；建议改为“{professional_errors[0]['fix']}”")
    elif debug.get("professional_score", 10) < 7:
        suggestions.append("补强专业校验与责任意识：避免绝对化表述，并把证据、机制与边界写清。")

    if debug.get("evidence_quality", 0) < 12:
        suggestions.append("增加清晰截图并圈出关键字段，提高证据可定位性。")
    if debug.get("expression_score", 0) < 7:
        suggestions.append("按“目标与边界 -> 任务 -> 关系图/表 -> 机制解释 -> 风险/防护 -> 自评”的顺序重组结构。")
    if content_issues and not suggestions:
        suggestions.append(f"优先处理：{content_issues[0]}。")
    if not suggestions:
        suggestions.append("进一步压缩抽象表述，把结论与具体截图、字段、URL、参数或端口一一对应。")
    return " ".join(suggestions[:2])


def label_for_score(score: int, hard_gate_reasons: list[str]) -> str:
    if "未提交本章作业" in hard_gate_reasons:
        return "需补交"
    if score >= 90:
        return "优秀"
    if score >= 80:
        return "良好"
    if score >= 70:
        return "中等"
    if score >= 60:
        return "及格"
    return "待补救"


def _load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def load_profile(path: Path) -> ChapterProfile:
    raw = _load_json(path)
    tasks = [TaskRule(**item) for item in raw.get("tasks", [])]
    redlines = [RedlineRule(**item) for item in raw.get("redlines", [])]
    return ChapterProfile(
        course_name=raw["course_name"],
        chapter_name=raw["chapter_name"],
        chapter_mainline=raw["chapter_mainline"],
        capability_goals=raw.get("capability_goals", []),
        tasks=tasks,
        relation_item_name=raw.get("relation_item_name", "关系图/关系表"),
        relation_item_score=float(raw.get("relation_item_score", 0)),
        self_eval_score=float(raw.get("self_eval_score", 0)),
        redlines=redlines,
        professional_checks=raw.get("professional_checks", []),
        default_tags=raw.get("default_tags", []),
    )


def validate_profile(profile: ChapterProfile) -> list[str]:
    errors: list[str] = []
    task_total = sum(task.score for task in profile.tasks)
    total = task_total + profile.relation_item_score + profile.self_eval_score
    if task_total <= 0:
        errors.append("任务分值总和必须大于 0。")
    if total > 100:
        errors.append(f"任务分值 + 收口项分值超过 100 分：当前为 {total}。")
    if not profile.capability_goals:
        errors.append("能力目标不能为空。")
    if not profile.redlines:
        errors.append("至少应定义 1 条红线规则。")
    return errors


def summarize_profile(profile: ChapterProfile) -> dict[str, Any]:
    return {
        "course_name": profile.course_name,
        "chapter_name": profile.chapter_name,
        "chapter_mainline": profile.chapter_mainline,
        "capability_goal_count": len(profile.capability_goals),
        "task_count": len(profile.tasks),
        "task_score_total": round(sum(task.score for task in profile.tasks), 2),
        "relation_item_name": profile.relation_item_name,
        "relation_item_score": profile.relation_item_score,
        "self_eval_score": profile.self_eval_score,
        "redline_count": len(profile.redlines),
        "default_tag_count": len(profile.default_tags),
    }


class OCRExtractor:
    def __init__(self) -> None:
        self.ocr = RapidOCR()

    def extract_pdf_text(self, pdf_path: Path) -> tuple[str, int]:
        doc = fitz.open(pdf_path)
        pages: list[str] = []
        for page in doc:
            text = normalize_text(page.get_text("text"))
            if len(text) < 120:
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                result, _ = self.ocr(pix.tobytes("png"))
                if result:
                    ocr_text = "\n".join(item[1] for item in result)
                    text = normalize_text(f"{text}\n{ocr_text}".strip())
            pages.append(text)
        page_count = doc.page_count
        doc.close()
        return "\n\n".join(pages).strip(), page_count


def build_heading_patterns(name: str, index: int) -> list[str]:
    clean = name.strip()
    task_num = str(index)
    cn_num = CN_NUMS[index - 1] if index - 1 < len(CN_NUMS) else task_num
    variants = [clean]
    for part in re.split(r"[/()（）·\s-]+", clean):
        part = part.strip()
        if len(part) >= 2 and part not in variants:
            variants.append(part)

    patterns: list[str] = []
    for variant in variants:
        escaped = re.escape(variant)
        patterns.extend(
            [
                rf"(?m)^\s*任务\s*[{task_num}{cn_num}].*{escaped}.*$",
                rf"(?m)^\s*(?:{task_num}|{cn_num})\s*[.．、:：)]\s*.*{escaped}.*$",
                rf"(?m)^\s*.*{escaped}.*$",
            ]
        )
    return patterns


def locate_sections_generic(text: str, profile: ChapterProfile) -> dict[str, str]:
    heading_specs: list[tuple[str, list[str]]] = []
    for index, task in enumerate(profile.tasks, start=1):
        heading_specs.append((f"task{index}", build_heading_patterns(task.name, index)))

    relation_patterns = [rf"(?m)^\s*.*{re.escape(name.strip())}.*$" for name in re.split(r"[/、]", profile.relation_item_name) if name.strip()]
    relation_patterns.extend(
        [
            r"(?m)^\s*.*综合关系图.*$",
            r"(?m)^\s*.*综合关系表.*$",
            r"(?m)^\s*.*因果链.*$",
            r"(?m)^\s*.*流程图.*$",
            r"(?m)^\s*.*关系链.*$",
        ]
    )
    heading_specs.extend(
        [
            ("relation", relation_patterns),
            ("boundary", [r"(?m)^\s*.*授权边界.*$", r"(?m)^\s*.*目标与边界.*$", r"(?m)^\s*.*授权与红线.*$"]),
            ("risk", [r"(?m)^\s*.*风险.*防护.*$", r"(?m)^\s*.*防护.*说明.*$", r"(?m)^\s*.*风险.*说明.*$"]),
            ("ai_review", [r"(?m)^\s*.*AI输出审核.*$", r"(?m)^\s*.*人工复核.*$", r"(?m)^\s*.*人工审核痕迹.*$"]),
            ("self_eval", [r"(?m)^\s*.*学生自评表.*$", r"(?m)^\s*.*自评表.*$"]),
        ]
    )

    positions: list[tuple[int, str]] = []
    for key, patterns in heading_specs:
        starts: list[int] = []
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                starts.append(match.start())
        if starts:
            positions.append((min(starts), key))

    positions.sort()
    sections: dict[str, str] = {}
    for index, (start, key) in enumerate(positions):
        end = positions[index + 1][0] if index + 1 < len(positions) else len(text)
        sections[key] = text[start:end].strip()
    return sections


def locate_sections_chapter2(text: str) -> dict[str, str]:
    cn_num_prefix = r"(?:[一二三四五六七八九十]+\s*[、.．：:]?\s*)?"
    heading_specs = [
        ("boundary", [r"(?m)^\s*(?:\d+\s*[.．、]?\s*)?(?:目标与授权边界|授权边界说明|授权边界)\s*$"]),
        ("task1", [rf"(?m)^\s*(?:任务\s*[1一]|{cn_num_prefix}Google\s*公开页面线索(?:记录|观察)?).*"]),
        ("task2", [rf"(?m)^\s*(?:任务\s*[2二]|{cn_num_prefix}Whois\s*域名身份信息(?:记录)?).*"]),
        ("task3", [rf"(?m)^\s*(?:任务\s*[3三]|{cn_num_prefix}Shodan\s*服务暴露线索(?:记录)?).*"]),
        ("task4", [rf"(?m)^\s*(?:任务\s*[4四]|{cn_num_prefix}社交工程风险场景分析).*"]),
        ("task5", [rf"(?m)^\s*(?:任务\s*[5五]|{cn_num_prefix}Burp\s*请求证据观察).*"]),
        ("task6", [rf"(?m)^\s*(?:任务\s*[6六]|{cn_num_prefix}Nmap\s*授权主动探测记录).*"]),
        ("relation", [r"(?m)^\s*(?:\d+\s*[.．、]?\s*)?(?:综合关系图|综合关系表|六层信息综合关系图|六层信息综合关系表)\s*$"]),
        ("mechanism", [r"(?m)^\s*(?:\d+\s*[.．、]?\s*)?机制解释(?:与工具说明)?\s*$"]),
        ("tool_mechanism", [r"(?m)^\s*(?:\d+\s*[.．、]?\s*)?(?:每个工具的最小机制说明|工具机制概述表|工具机制速填表|工具说明)\s*$"]),
        ("risk", [r"(?m)^\s*(?:\d+\s*[.．、]?\s*)?(?:风险\s*/\s*防护说明|风险或防护说明|风险与防护说明|风险说明)\s*$"]),
        ("ai_review", [r"(?m)^\s*(?:\d+\s*[.．、]?\s*)?(?:人工审核痕迹(?:（必填）)?|人工复核痕迹|AI输出审核与人工复核记录(?:（必填）)?|AI输出审核(?:与人工复核记录)?(?:（必填）)?)\s*$"]),
        ("self_eval", [r"(?m)^\s*(?:\d+\s*[.．、]?\s*)?(?:学生自评表|自评表)\s*$"]),
    ]
    positions: list[tuple[int, str]] = []
    for key, patterns in heading_specs:
        starts: list[int] = []
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                starts.append(match.start())
        if starts:
            positions.append((min(starts), key))
    positions.sort()
    sections: dict[str, str] = {}
    for index, (start, key) in enumerate(positions):
        end = positions[index + 1][0] if index + 1 < len(positions) else len(text)
        sections[key] = text[start:end].strip()
    return sections


def locate_sections_for_profile(text: str, profile: ChapterProfile) -> dict[str, str]:
    if "第二章" in profile.chapter_name:
        return locate_sections_chapter2(text)
    return locate_sections_generic(text, profile)


def evaluate_relation_section(section_text: str, signal_terms: list[str], max_score: float) -> dict[str, Any]:
    submitted = bool(section_text)
    signal_hits = sum(1 for term in signal_terms if term and term in section_text)
    relation_hits = sum(
        [
            contains_any(section_text, ["关系", "关系链", "流程", "链路", "因果"]),
            contains_any(section_text, ["控制点", "风险点", "防护"]),
            contains_any(section_text, ["下一步", "结果类型", "能说明什么", "不能替代什么"]),
            contains_any(section_text, ["->", "→", "=>", "⇒"]),
        ]
    )
    if submitted and signal_hits >= 5 and relation_hits >= 3:
        return {"submitted": True, "status": "complete", "score": max_score, "signal_hits": signal_hits, "relation_hits": relation_hits}
    if submitted:
        return {"submitted": True, "status": "partial", "score": round(max_score * 0.5, 2), "signal_hits": signal_hits, "relation_hits": relation_hits}
    return {"submitted": False, "status": "missing", "score": 0.0, "signal_hits": 0, "relation_hits": 0}


def evaluate_graph_section_chapter2(normalized: str, graph_section: str, max_score: float) -> dict[str, Any]:
    graph_markers = ["综合关系图", "综合关系表", "六层信息综合关系图", "六层信息综合关系表"]
    graph_submitted = bool(graph_section) or contains_any(normalized, graph_markers)

    header_hits = sum(
        [
            contains_any(graph_section, ["信息层"]),
            contains_any(graph_section, ["工具/方法", "工具", "方法"]),
            contains_any(graph_section, ["结果类型"]),
            contains_any(graph_section, ["能说明什么"]),
            contains_any(graph_section, ["不能替代什么", "不能替代"]),
            contains_any(graph_section, ["下一步"]),
        ]
    )
    layer_hits = compact_contains(
        graph_section,
        [
            "页面线索层",
            "域名身份层",
            "域名身份线索层",
            "服务索引层",
            "服务索引线索层",
            "人员组织层",
            "人员与组织信息层",
            "请求证据层",
            "授权探测层",
            "授权主动探测层",
        ],
    )
    relation_hits = sum(
        [
            contains_any(graph_section, ["结果类型"]),
            contains_any(graph_section, ["能说明什么"]),
            contains_any(graph_section, ["不能替代什么", "不能替代"]),
            contains_any(graph_section, ["下一步"]),
        ]
    )

    if graph_submitted and header_hits >= 4 and layer_hits >= 3 and relation_hits >= 3:
        status = "complete"
        score = max_score
    elif graph_submitted:
        status = "partial"
        score = round(max_score * 0.5, 2)
    else:
        status = "missing"
        score = 0.0

    return {
        "submitted": graph_submitted,
        "status": status,
        "score": score,
        "header_hits": header_hits,
        "layer_hits": layer_hits,
        "relation_hits": relation_hits,
    }


def build_similarity_profile(text: str, profile: ChapterProfile) -> dict[str, Any]:
    normalized = normalize_text(text)
    sections = locate_sections_for_profile(normalized, profile)
    payload: dict[str, str] = {}
    for index, _task in enumerate(profile.tasks, start=1):
        key = f"task{index}"
        section_text = sections.get(key, "")
        cleaned = clean_similarity_text(section_text)
        if len(cleaned) >= 60:
            payload[key] = cleaned
    for key in ["relation", "mechanism", "risk", "ai_review"]:
        section_text = sections.get(key, "")
        cleaned = clean_similarity_text(section_text)
        if len(cleaned) >= 60:
            payload[key] = cleaned
    return {"full": clean_similarity_text(normalized), "sections": payload}


def text_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def longest_block_ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    matcher = SequenceMatcher(None, a, b)
    longest = matcher.find_longest_match(0, len(a), 0, len(b)).size
    return longest / max(1, min(len(a), len(b)))


def detect_similarity_reviews(profile: ChapterProfile, submission_texts: dict[str, str]) -> dict[str, dict[str, Any]]:
    profiles = {name: build_similarity_profile(text, profile) for name, text in submission_texts.items()}
    reviews: dict[str, dict[str, Any]] = {
        name: {
            "is_suspected_similar": False,
            "is_highly_similar": False,
            "max_full_ratio": 0.0,
            "max_longest_block_ratio": 0.0,
            "matched_peers": [],
        }
        for name in submission_texts
    }
    names = sorted(submission_texts)
    for index, name_a in enumerate(names):
        for name_b in names[index + 1 :]:
            profile_a = profiles[name_a]
            profile_b = profiles[name_b]
            full_ratio = text_similarity(profile_a["full"], profile_b["full"])
            block_ratio = longest_block_ratio(profile_a["full"], profile_b["full"])

            common_keys = set(profile_a["sections"]).intersection(profile_b["sections"])
            section_scores: dict[str, float] = {}
            high_sections: list[str] = []
            very_high_sections: list[str] = []
            for key in common_keys:
                ratio = text_similarity(profile_a["sections"][key], profile_b["sections"][key])
                section_scores[key] = round(ratio, 3)
                if ratio >= 0.9:
                    high_sections.append(key)
                if ratio >= 0.96:
                    very_high_sections.append(key)

            is_suspected = full_ratio >= 0.88 and len(high_sections) >= max(3, len(common_keys) // 2)
            is_highly_similar = (
                (full_ratio >= 0.93 and len(high_sections) >= max(4, len(common_keys) // 2 + 1))
                or (block_ratio >= 0.6 and len(very_high_sections) >= 3 and len(high_sections) >= 4)
            )
            if not (is_suspected or is_highly_similar):
                continue

            peer_summary = {
                "peer": name_b,
                "full_ratio": round(full_ratio, 3),
                "longest_block_ratio": round(block_ratio, 3),
                "high_sections": sorted(high_sections),
                "very_high_sections": sorted(very_high_sections),
                "section_scores": {key: section_scores[key] for key in sorted(section_scores)},
            }
            reviews[name_a]["matched_peers"].append(peer_summary)
            reviews[name_a]["max_full_ratio"] = max(reviews[name_a]["max_full_ratio"], full_ratio)
            reviews[name_a]["max_longest_block_ratio"] = max(reviews[name_a]["max_longest_block_ratio"], block_ratio)
            reviews[name_a]["is_suspected_similar"] = reviews[name_a]["is_suspected_similar"] or is_suspected
            reviews[name_a]["is_highly_similar"] = reviews[name_a]["is_highly_similar"] or is_highly_similar

            reverse_summary = dict(peer_summary)
            reverse_summary["peer"] = name_a
            reviews[name_b]["matched_peers"].append(reverse_summary)
            reviews[name_b]["max_full_ratio"] = max(reviews[name_b]["max_full_ratio"], full_ratio)
            reviews[name_b]["max_longest_block_ratio"] = max(reviews[name_b]["max_longest_block_ratio"], block_ratio)
            reviews[name_b]["is_suspected_similar"] = reviews[name_b]["is_suspected_similar"] or is_suspected
            reviews[name_b]["is_highly_similar"] = reviews[name_b]["is_highly_similar"] or is_highly_similar

    for review in reviews.values():
        review["max_full_ratio"] = round(review["max_full_ratio"], 3)
        review["max_longest_block_ratio"] = round(review["max_longest_block_ratio"], 3)
        review["matched_peers"] = sorted(
            review["matched_peers"],
            key=lambda item: (item["full_ratio"], item["longest_block_ratio"], len(item["high_sections"])),
            reverse=True,
        )[:3]
    return reviews


def build_professional_review(
    normalized: str,
    ai_review_trace: dict[str, Any],
    checks: list[dict[str, Any]],
) -> dict[str, Any]:
    ai_like_paragraphs = detect_ai_like_paragraphs(normalized)
    professional_errors: list[dict[str, str]] = []
    unchecked_traces: list[str] = []
    score = 10.0

    for item in checks:
        if item["ok"]:
            continue
        professional_errors.append(
            {
                "item": item["item"],
                "why": item["why"],
                "fix": item["fix"],
            }
        )
        score -= float(item.get("penalty", 2.0))

    absolute_hits = [phrase for phrase in ABSOLUTE_ASSERTIONS if phrase in normalized]
    if absolute_hits:
        professional_errors.append(
            {
                "item": f"存在未经核实的绝对化表述：{absolute_hits[0]}",
                "why": "这类表述直接把线索、观察结果或实验现象写成确定结论，但文中没有相应验证过程，会造成专业误判。",
                "fix": "改成“提示可能存在风险”“当前现象说明需要进一步验证”“该结果不足以直接证明漏洞成立”。",
            }
        )
        unchecked_traces.append(f"出现绝对化表述“{absolute_hits[0]}”，但缺少相应验证或限定条件。")
        score -= 2.0

    if not ai_review_trace["present"]:
        unchecked_traces.append("未明显看到 AI输出审核与人工复核记录，无法判断是否审核过 AI 输出。")
        score -= 1.5
    elif not ai_review_trace["complete"]:
        unchecked_traces.append("AI输出审核与人工复核记录已出现，但删改、核实或边界说明仍不完整。")
        score -= 0.7

    if ai_like_paragraphs:
        unchecked_traces.append("部分段落语言较模板化，抽象概括较多，但与具体证据绑定较弱。")
        score -= 0.5

    if professional_errors and not unchecked_traces:
        unchecked_traces.append("文中存在专业判断与证据绑定不足或概念层次混淆，疑似提交前未做充分审核。")

    return {
        "score": round(clamp(score, 0.0, 10.0), 2),
        "ai_like_paragraphs": ai_like_paragraphs[:2],
        "unchecked_traces": unchecked_traces[:2],
        "professional_errors": professional_errors[:3],
        "absolute_hits": absolute_hits[:2],
        "ai_review_trace": ai_review_trace,
    }


def detect_public_targets(text: str) -> list[str]:
    public_targets: list[str] = []
    for candidate in extract_domains(text):
        lowered = candidate.lower()
        if not is_private_target(lowered) and "dvwa" not in lowered and "localhost" not in lowered:
            public_targets.append(candidate)
    for ip in IP_RE.findall(text):
        if not is_private_target(ip):
            public_targets.append(ip)
    unique_targets: list[str] = []
    for item in public_targets:
        if item not in unique_targets:
            unique_targets.append(item)
    return unique_targets


def finalize_specialized_scoring(
    *,
    normalized: str,
    page_count: int,
    profile: ChapterProfile,
    sections: dict[str, str],
    task_scores: dict[str, float],
    relation_eval: dict[str, Any],
    self_eval_present: bool,
    ai_review_trace: dict[str, Any],
    evidence_quality: float,
    mechanism_score: float,
    risk_score: float,
    protection_score: float,
    expression_score: float,
    professional_review: dict[str, Any],
    similarity_review: dict[str, Any] | None,
    rubric_used: str,
    extra_content_issues: list[str] | None = None,
    extra_hard_gate_reasons: list[str] | None = None,
    extra_debug: dict[str, Any] | None = None,
) -> tuple[int, dict[str, Any]]:
    task_completion = clamp(
        sum(task_scores.values()) + relation_eval["score"] + (profile.self_eval_score if self_eval_present else 0.0),
        0.0,
        20.0,
    )
    professional_score = professional_review["score"]
    base_score = task_completion + evidence_quality + mechanism_score + risk_score + protection_score + expression_score + professional_score

    content_issues: list[str] = []
    for key, value in task_scores.items():
        max_score = next(task.score for index, task in enumerate(profile.tasks, start=1) if f"task{index}" == key)
        if value < 0.6 * max_score:
            content_issues.append(f"{profile.tasks[int(key[-1]) - 1].name}内容不全")
    if relation_eval["status"] == "missing":
        content_issues.append(f"未提交{profile.relation_item_name}")
    elif relation_eval["status"] == "partial":
        content_issues.append(f"{profile.relation_item_name}已提交但关系表达不足")
    if not self_eval_present:
        content_issues.append("缺少学生自评表")
    if not ai_review_trace["present"]:
        content_issues.append("缺少AI输出审核与人工复核记录")
    if professional_score < 7 and professional_review["professional_errors"]:
        content_issues.append("专业校验与责任意识不足")
    if similarity_review and similarity_review.get("is_highly_similar"):
        content_issues.append("与同班作业高度雷同")
    elif similarity_review and similarity_review.get("is_suspected_similar"):
        content_issues.append("存在作业雷同风险，建议人工复核")
    if extra_content_issues:
        for issue in extra_content_issues:
            if issue not in content_issues:
                content_issues.append(issue)

    hard_gate_reasons: list[str] = []
    if SESSION_LEAK_RE.search(normalized) and not MASK_HINT_RE.search(normalized):
        hard_gate_reasons.append("敏感信息疑似未脱敏")
    if similarity_review and similarity_review.get("is_highly_similar"):
        hard_gate_reasons.append("与同班作业高度雷同")
    if extra_hard_gate_reasons:
        for item in extra_hard_gate_reasons:
            if item not in hard_gate_reasons:
                hard_gate_reasons.append(item)

    final_score = int(round(clamp(base_score, 0.0, 100.0)))
    if hard_gate_reasons:
        final_score = min(final_score, 59)

    debug = {
        "rubric_used": rubric_used,
        "task_completion": round(task_completion, 2),
        "evidence_quality": round(evidence_quality, 2),
        "mechanism_score": round(mechanism_score, 2),
        "risk_score": round(risk_score, 2),
        "protection_score": round(protection_score, 2),
        "expression_score": round(expression_score, 2),
        "professional_score": round(professional_score, 2),
        "base_score": round(base_score, 2),
        "final_score": final_score,
        "page_count": page_count,
        "text_length": len(normalized),
        "relation_status": relation_eval["status"],
        "self_eval_present": self_eval_present,
        "ai_review_present": ai_review_trace["present"],
        "ai_review_complete": ai_review_trace["complete"],
        "task_scores": {key: round(value, 2) for key, value in task_scores.items()},
        "task_max_scores": {f"task{index}": task.score for index, task in enumerate(profile.tasks, start=1)},
        "task_name_map": {f"task{index}": task.name for index, task in enumerate(profile.tasks, start=1)},
        "task_requirement_map": {f"task{index}": task.semantic_requirements for index, task in enumerate(profile.tasks, start=1)},
        "content_issues": content_issues,
        "hard_gate_reasons": hard_gate_reasons,
        "ai_like_paragraphs": professional_review["ai_like_paragraphs"],
        "unchecked_traces": professional_review["unchecked_traces"],
        "professional_errors": professional_review["professional_errors"],
        "absolute_hits": professional_review["absolute_hits"],
        "similarity_review": similarity_review or {
            "is_suspected_similar": False,
            "is_highly_similar": False,
            "max_full_ratio": 0.0,
            "max_longest_block_ratio": 0.0,
            "matched_peers": [],
        },
    }
    if extra_debug:
        debug.update(extra_debug)
    return final_score, debug


def chapter1_scoring(text: str, page_count: int, profile: ChapterProfile, similarity_review: dict[str, Any] | None) -> tuple[int, dict[str, Any]]:
    normalized = normalize_text(text)
    lower = normalized.lower()
    sections = locate_sections_generic(normalized, profile)

    task1 = sections.get("task1", "") or find_context_window(normalized, ["浏览器", "服务器", "数据库", "三层架构"])
    task2 = sections.get("task2", "") or find_context_window(normalized, ["HTTP", "HTTPS", "请求", "响应", "状态码"])
    task3 = sections.get("task3", "") or find_context_window(normalized, ["Cookie", "Session", "登录状态", "状态维持"])
    task4 = sections.get("task4", "") or find_context_window(normalized, ["输入—处理—输出—执行—身份", "输入-处理-输出-执行-身份", "风险点", "控制点"])
    relation_section = sections.get("relation", "") or find_context_window(normalized, ["关系图", "链路图", "流程图", "浏览器", "服务器", "数据库"])
    boundary_section = sections.get("boundary", "") or find_context_window(normalized, ["授权边界", "授权", "允许", "禁止", "脱敏"])
    risk_section = sections.get("risk", "") or find_context_window(normalized, ["风险", "控制点", "防护", "安全关注点"])
    ai_review_section = sections.get("ai_review", "") or find_context_window(normalized, ["AI输出审核", "人工复核", "人工审核痕迹"])
    self_eval_section = sections.get("self_eval", "") or find_context_window(normalized, ["学生自评表", "自评表"])

    t1_checks = [
        contains_any(task1, ["浏览器"]) and contains_any(task1, ["服务器"]) and contains_any(task1, ["数据库"]),
        contains_any(task1, ["输入"]) or contains_any(task1, ["用户输入"]),
        contains_any(task1, ["返回", "结果", "响应"]),
        contains_any(task1, ["风险点", "控制点", "安全关注点"]),
    ]
    t1_global = [
        contains_any(normalized, ["浏览器"]) and contains_any(normalized, ["服务器"]) and contains_any(normalized, ["数据库"]),
        contains_any(normalized, ["输入"]) or contains_any(normalized, ["用户输入"]),
        contains_any(normalized, ["返回", "结果", "响应"]),
        contains_any(normalized, ["风险点", "控制点", "安全关注点"]),
    ]

    t2_checks = [
        contains_any(task2, ["请求方法", "Method", "GET", "POST"]),
        contains_any(task2, ["URL", "路径", "Request URL", "请求URL"]),
        contains_any(task2, ["参数", "参数位置", "表单参数", "URL 参数"]),
        contains_any(task2, ["状态码", "Status", "返回结果", "响应"]),
        contains_any(task2, ["HTTPS"]) and contains_any(task2, ["传输", "加密", "业务安全", "逻辑问题"]),
    ]
    t2_global = [
        contains_any(normalized, ["请求方法", "Method", "GET", "POST"]),
        contains_any(normalized, ["URL", "路径", "Request URL", "请求URL"]),
        contains_any(normalized, ["参数", "参数位置", "表单参数", "URL 参数"]),
        contains_any(normalized, ["状态码", "Status", "返回结果", "响应"]),
        contains_any(normalized, ["HTTPS"]) and contains_any(normalized, ["传输", "加密", "业务安全", "逻辑问题"]),
    ]

    t3_checks = [
        contains_any(task3, ["登录前", "登录后", "状态变化", "已登录"]),
        contains_any(task3, ["Cookie"]),
        contains_any(task3, ["Session"]),
        contains_any(task3, ["为什么登录后还能保持状态", "保持状态", "状态维持", "无状态"]),
        contains_any(task3, ["浏览器侧", "服务器侧", "分工", "职责"]),
    ]
    t3_global = [
        contains_any(normalized, ["登录前", "登录后", "状态变化", "已登录"]),
        contains_any(normalized, ["Cookie"]),
        contains_any(normalized, ["Session"]),
        contains_any(normalized, ["为什么登录后还能保持状态", "保持状态", "状态维持", "无状态"]),
        contains_any(normalized, ["浏览器侧", "服务器侧", "分工", "职责"]),
    ]

    t4_checks = [
        contains_any(task4, ["输入"]) and contains_any(task4, ["处理"]) and contains_any(task4, ["输出"]) and contains_any(task4, ["身份"]),
        contains_any(task4, ["页面场景", "场景", "页面"]),
        contains_any(task4, ["风险点"]) or count_hits(task4, ["风险点", "控制点"]) >= 2,
        contains_any(task4, ["统一框架", "分析框架", "输入—处理—输出—执行—身份", "输入-处理-输出-执行-身份"]),
    ]
    t4_global = [
        contains_any(normalized, ["输入"]) and contains_any(normalized, ["处理"]) and contains_any(normalized, ["输出"]) and contains_any(normalized, ["身份"]),
        contains_any(normalized, ["页面场景", "场景", "页面"]),
        contains_any(normalized, ["风险点"]) or count_hits(normalized, ["风险点", "控制点"]) >= 2,
        contains_any(normalized, ["统一框架", "分析框架", "输入—处理—输出—执行—身份", "输入-处理-输出-执行-身份"]),
    ]

    task_scores = {
        "task1": max_ratio_score(t1_checks, t1_global, profile.tasks[0].score),
        "task2": max_ratio_score(t2_checks, t2_global, profile.tasks[1].score),
        "task3": max_ratio_score(t3_checks, t3_global, profile.tasks[2].score),
        "task4": max_ratio_score(t4_checks, t4_global, profile.tasks[3].score),
    }

    relation_eval = evaluate_relation_section(
        relation_section,
        ["浏览器", "服务器", "数据库", "请求", "响应", "Cookie", "Session", "输入", "处理", "输出", "身份"],
        profile.relation_item_score,
    )
    self_eval_present = bool(self_eval_section) and contains_any(self_eval_section, ["任务完成情况", "机制理解", "证据质量", "自评结论"])
    ai_review_trace = detect_ai_review_trace(ai_review_section or normalized)

    task_completion = sum(task_scores.values()) + relation_eval["score"] + (profile.self_eval_score if self_eval_present else 0.0)
    task_completion = clamp(task_completion, 0.0, 20.0)

    screenshot_markers = count_regex(normalized, SCREENSHOT_MARKER_RE)
    evidence_categories = sum(1 for value in task_scores.values() if value >= 0.7 * profile.tasks[list(task_scores).index(next(k for k,v in task_scores.items() if v==value))].score) if False else 0
    evidence_categories = sum(
        [
            1 if task_scores["task1"] >= 2.8 else 0,
            1 if task_scores["task2"] >= 2.8 else 0,
            1 if task_scores["task3"] >= 2.8 else 0,
            1 if task_scores["task4"] >= 2.0 else 0,
        ]
    )
    artifact_categories = sum(
        [
            1 if bool(URL_RE.search(normalized)) else 0,
            1 if contains_any(normalized, ["GET", "POST", "状态码", "HTTP", "HTTPS"]) else 0,
            1 if contains_any(normalized, ["Cookie", "Session"]) else 0,
            1 if contains_any(normalized, ["浏览器", "服务器", "数据库"]) else 0,
            1 if contains_any(normalized, ["输入", "处理", "输出", "身份"]) else 0,
        ]
    )
    evidence_quality = clamp(
        page_points(page_count)
        + screenshot_points(screenshot_markers)
        + ratio_score(evidence_categories, 4, 4.5)
        + ratio_score(artifact_categories, 5, 4.5),
        0.0,
        18.0,
    )

    mechanism_checks = [
        contains_any(normalized, ["请求", "响应"]),
        contains_any(normalized, ["无状态", "状态维持"]),
        contains_any(normalized, ["Cookie"]) and contains_any(normalized, ["Session"]),
        contains_any(normalized, ["浏览器", "服务器", "数据库"]),
        contains_any(normalized, ["输入—处理—输出—执行—身份", "输入-处理-输出-执行-身份"]),
    ]
    mechanism_score = ratio_score(sum(mechanism_checks), len(mechanism_checks), 20.0)

    risk_checks = [
        contains_any(boundary_section or normalized, ["授权", "边界", "允许", "禁止"]),
        contains_any(boundary_section or normalized, ["脱敏"]),
        contains_any(task4 + "\n" + risk_section, ["风险点", "控制点", "安全关注点"]),
    ]
    risk_score = ratio_score(sum(risk_checks), len(risk_checks), 12.0)

    control_hits = count_hits(
        normalized,
        ["控制点", "HTTPS", "HttpOnly", "Secure", "脱敏", "授权", "限制", "验证", "边界"],
    )
    protection_score = 10.0 if control_hits >= 5 else (8.0 if control_hits >= 4 else (6.0 if control_hits >= 3 else (3.0 if control_hits >= 2 else 0.0)))

    ordered_sections = all(bool(sections.get(f"task{index}")) for index in range(1, len(profile.tasks) + 1))
    structure_points = 5.0 if ordered_sections and relation_eval["status"] == "complete" else (4.0 if ordered_sections else 2.5)
    text_length = len(normalized)
    expression_points = 5.0 if text_length >= 2200 else (4.0 if text_length >= 1600 else (3.0 if text_length >= 1000 else 1.5))
    expression_score = clamp(structure_points + expression_points, 0.0, 10.0)

    ip_as_identity = contains_any(normalized, ["依赖ip地址", "主要依赖ip", "靠ip识别"])
    ip_mitigated = contains_any(normalized, ["不是主要依据", "不是主要身份依据", "Cookie", "Session"])
    boundary_good = contains_any(boundary_section or normalized, ["授权", "脱敏", "边界"])
    professional_review = build_professional_review(
        normalized,
        ai_review_trace,
        checks=[
            {
                "ok": contains_any(normalized, ["HTTPS"]) and contains_any(normalized, ["传输", "加密", "业务安全", "逻辑问题"]),
                "item": "HTTP/HTTPS 边界说明不足",
                "why": "第一章要建立“HTTPS 保护传输，不等于应用逻辑天然安全”的最小边界。",
                "fix": "补写 HTTPS 主要保护传输过程，不能直接替代对输入处理、输出和身份控制的安全分析。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(normalized, ["Cookie"]) and contains_any(normalized, ["Session"]) and contains_any(normalized, ["分工", "职责", "浏览器侧", "服务器侧"]),
                "item": "Cookie 与 Session 分工说明不足",
                "why": "多数学生会记名词，但不会把浏览器侧和服务器侧状态维持机制连成因果链。",
                "fix": "明确写出：Cookie 保存并回传状态线索，Session 保存服务器侧会话记录，二者配合实现状态维持。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(task3 + "\n" + normalized, ["登录前", "登录后", "状态变化", "换浏览器", "Cookie丢失", "保持状态"]),
                "item": "状态维持解释缺少现象证据",
                "why": "第一章不只是记住 Cookie/Session 术语，而是要能回到登录前后状态变化解释为什么系统能识别用户。",
                "fix": "补写登录前后状态变化、换浏览器或丢失 Cookie 时的现象，并把它与状态维持机制对应起来。",
                "penalty": 2.0,
            },
            {
                "ok": not (ip_as_identity and not ip_mitigated),
                "item": "把 IP 地址绝对化写成主要身份识别依据",
                "why": "第一章应避免把 IP 误写成主要身份识别基础，否则会直接破坏后续会话安全章节的理解。",
                "fix": "改写为：IP 可能作为环境信息，但登录状态持续识别主要依赖 Cookie 与 Session 等状态维持机制。",
                "penalty": 2.0,
            },
            {
                "ok": boundary_good,
                "item": "授权边界与脱敏表达不足",
                "why": "第一章首先建立边界意识。如果连授权和脱敏都没写清，后续所有技术观察都会失去责任边界。",
                "fix": "补写允许环境、禁止行为和脱敏要求，明确实验仅限 DVWA、本机或教师授权环境。",
                "penalty": 2.0,
            },
        ],
    )
    professional_score = professional_review["score"]

    base_score = task_completion + evidence_quality + mechanism_score + risk_score + protection_score + expression_score + professional_score

    content_issues: list[str] = []
    for key, value in task_scores.items():
        max_score = next(task.score for index, task in enumerate(profile.tasks, start=1) if f"task{index}" == key)
        if value < 0.6 * max_score:
            content_issues.append(f"{profile.tasks[int(key[-1]) - 1].name}内容不全")
    if relation_eval["status"] == "missing":
        content_issues.append(f"未提交{profile.relation_item_name}")
    elif relation_eval["status"] == "partial":
        content_issues.append(f"{profile.relation_item_name}已提交但关系表达不足")
    if not self_eval_present:
        content_issues.append("缺少学生自评表")
    if not ai_review_trace["present"]:
        content_issues.append("缺少AI输出审核与人工复核记录")
    if professional_score < 7 and professional_review["professional_errors"]:
        content_issues.append("专业校验与责任意识不足")
    if similarity_review and similarity_review.get("is_highly_similar"):
        content_issues.append("与同班作业高度雷同")
    elif similarity_review and similarity_review.get("is_suspected_similar"):
        content_issues.append("存在作业雷同风险，建议人工复核")

    hard_gate_reasons: list[str] = []
    if SESSION_LEAK_RE.search(normalized) and not MASK_HINT_RE.search(normalized):
        hard_gate_reasons.append("敏感信息疑似未脱敏")
    if similarity_review and similarity_review.get("is_highly_similar"):
        hard_gate_reasons.append("与同班作业高度雷同")

    final_score = int(round(clamp(base_score, 0.0, 100.0)))
    if hard_gate_reasons:
        final_score = min(final_score, 59)

    debug = {
        "rubric_used": "projects5_v5_generic_chapter01",
        "task_completion": round(task_completion, 2),
        "evidence_quality": round(evidence_quality, 2),
        "mechanism_score": round(mechanism_score, 2),
        "risk_score": round(risk_score, 2),
        "protection_score": round(protection_score, 2),
        "expression_score": round(expression_score, 2),
        "professional_score": round(professional_score, 2),
        "base_score": round(base_score, 2),
        "final_score": final_score,
        "page_count": page_count,
        "text_length": text_length,
        "relation_status": relation_eval["status"],
        "relation_signal_hits": relation_eval["signal_hits"],
        "relation_relation_hits": relation_eval["relation_hits"],
        "self_eval_present": self_eval_present,
        "ai_review_present": ai_review_trace["present"],
        "ai_review_complete": ai_review_trace["complete"],
        "task_scores": {key: round(value, 2) for key, value in task_scores.items()},
        "task_max_scores": {f"task{index}": task.score for index, task in enumerate(profile.tasks, start=1)},
        "task_name_map": {f"task{index}": task.name for index, task in enumerate(profile.tasks, start=1)},
        "task_requirement_map": {f"task{index}": task.semantic_requirements for index, task in enumerate(profile.tasks, start=1)},
        "screenshot_markers": screenshot_markers,
        "content_issues": content_issues,
        "hard_gate_reasons": hard_gate_reasons,
        "domains": extract_domains(normalized)[:8],
        "ai_like_paragraphs": professional_review["ai_like_paragraphs"],
        "unchecked_traces": professional_review["unchecked_traces"],
        "professional_errors": professional_review["professional_errors"],
        "absolute_hits": professional_review["absolute_hits"],
        "similarity_review": similarity_review or {
            "is_suspected_similar": False,
            "is_highly_similar": False,
            "max_full_ratio": 0.0,
            "max_longest_block_ratio": 0.0,
            "matched_peers": [],
        },
    }
    return final_score, debug


def chapter2_professional_review(
    normalized: str,
    boundary_section: str,
    task3: str,
    task5: str,
    task6: str,
    nmap_not_vuln: bool,
    burp_evidence_reason: bool,
    ai_review_trace: dict[str, Any],
) -> dict[str, Any]:
    distinction_hit_count = sum(
        [
            contains_any(normalized, ["线索", "线索层"]),
            contains_any(normalized, ["证据", "技术证据", "请求证据"]),
            contains_any(normalized, ["主动探测", "主动探测结果"]),
            contains_any(normalized, ["第三方索引", "索引线索"]),
        ]
    )
    boundary_good = bool(boundary_section) and contains_any(boundary_section, ["授权", "允许", "禁止", "脱敏", "边界"])
    return build_professional_review(
        normalized,
        ai_review_trace,
        checks=[
            {
                "ok": distinction_hit_count >= 3,
                "item": "未充分区分线索、证据与主动探测",
                "why": "第二章核心不是工具罗列，而是区分公开线索、第三方索引、真实请求证据和授权主动探测。",
                "fix": "明确写出：Google/Whois/Shodan 属于线索层，Burp 属于请求证据层，Nmap 属于授权主动探测层。",
                "penalty": 3.0,
            },
            {
                "ok": not task3 or contains_any(task3, ["第三方索引", "历史索引", "不能替代授权扫描"]),
                "item": "Shodan 的结果类型说明不足",
                "why": "Shodan 返回的是第三方历史索引结果，不是你对目标即时做出的主动探测。",
                "fix": "补写“Shodan 属于第三方索引线索，不能替代授权扫描，也不能直接证明目标当前状态”。",
                "penalty": 1.0,
            },
            {
                "ok": not task5 or burp_evidence_reason,
                "item": "Burp 与线索层的证据差异说明不足",
                "why": "Burp 看到的是实时 HTTP 交互，证据价值高于公开线索和第三方索引。",
                "fix": "明确写出 Burp 捕获的是实时 HTTP 请求与响应，可看到 Method、URL、参数、Header、状态码，因此更接近技术证据。",
                "penalty": 1.0,
            },
            {
                "ok": not task6 or nmap_not_vuln,
                "item": "未明确说明开放端口不等于漏洞成立",
                "why": "Nmap 只能说明端口状态和可能服务，不能直接证明存在可利用漏洞。",
                "fix": "补写：开放端口或识别到服务，只表示攻击面存在，是否为漏洞仍需进一步验证。",
                "penalty": 2.0,
            },
            {
                "ok": boundary_good,
                "item": "授权边界与风险表达不足",
                "why": "第二章要求在授权前提下完成 Burp 和 Nmap，并明确禁止未授权抓包、扫描和社工接触。",
                "fix": "补写允许范围、禁止行为、脱敏要求，并说明 Burp/Nmap 仅限 DVWA、127.0.0.1 或教师明确授权目标。",
                "penalty": 2.0,
            },
        ],
    )


def chapter2_scoring(text: str, page_count: int, profile: ChapterProfile, similarity_review: dict[str, Any] | None) -> tuple[int, dict[str, Any]]:
    normalized = normalize_text(text)
    lower = normalized.lower()
    sections = locate_sections_chapter2(normalized)

    task1 = sections.get("task1", "") or find_context_window(normalized, ["Google 公开页面线索", "查询语句", "公开线索"])
    task2 = sections.get("task2", "") or find_context_window(normalized, ["Whois 域名身份信息", "Registrar", "Name Server"])
    task3 = sections.get("task3", "") or find_context_window(normalized, ["Shodan 服务暴露线索", "第三方索引", "Banner"])
    task4 = sections.get("task4", "") or find_context_window(normalized, ["社交工程风险场景分析", "案例名称", "防范建议"])
    task5 = sections.get("task5", "") or find_context_window(normalized, ["Burp 请求证据观察", "抓包目标", "Method", "Header"])
    task6 = sections.get("task6", "") or find_context_window(normalized, ["Nmap 授权主动探测记录", "扫描目标", "扫描命令", "开放端口"])
    relation_section = sections.get("relation", "") or find_context_window(
        normalized,
        ["综合关系图", "综合关系表", "六层信息综合关系表", "信息层", "下一步"],
    )
    mechanism_section = sections.get("mechanism", "") or find_context_window(normalized, ["机制解释", "被动观察", "主动探测"])
    tool_mechanism_section = sections.get("tool_mechanism", "") or find_context_window(normalized, ["每个工具的最小机制说明", "Google：", "Whois：", "Shodan：", "Burp", "Nmap"])
    boundary_section = sections.get("boundary", "") or find_context_window(normalized, ["目标与授权边界", "授权边界说明", "允许范围", "禁止行为"])
    risk_section = sections.get("risk", "") or find_context_window(normalized, ["风险与防护说明", "风险说明", "防护措施", "被动优先"])
    ai_review_section = sections.get("ai_review", "") or find_context_window(normalized, ["AI输出审核", "人工复核", "人工审核痕迹"])
    self_eval_section = sections.get("self_eval", "") or find_context_window(normalized, ["学生自评表", "任务完成情况", "自评结论"])

    t1_checks = [
        contains_any(task1, ["查询语句"]),
        contains_any(task1, ["site:", "intitle:", "inurl:", "filetype:", "语法元素"]),
        contains_any(task1, ["命中的结果类型", "结果类型"]),
        contains_any(task1, ["公开线索"]),
        contains_any(task1, ["只能算线索", "不能直接算证据", "不能直接作为技术证据", "不能直接作为证据"]),
        contains_any(task1, ["最小机制说明"]),
    ]
    t1_global_checks = [
        contains_any(normalized, ["目标名称"]) and contains_any(normalized, ["查询语句"]),
        contains_any(normalized, ["site:", "intitle:", "inurl:", "filetype:"]),
        contains_any(normalized, ["命中的结果类型", "结果类型"]),
        contains_any(normalized, ["公开线索"]),
        contains_any(normalized, ["只能算线索", "不能直接算证据", "不能直接作为技术证据", "不能直接作为证据"]),
        contains_any(normalized, ["最小机制说明"]) and contains_any(normalized, ["site:", "intitle:", "inurl:", "filetype:"]),
    ]
    t2_checks = [
        contains_any(task2, ["查询对象"]),
        sum([contains_any(task2, ["Registrar", "注册商"]), contains_any(task2, ["Creation Date", "注册时间"]), contains_any(task2, ["Expiration Date", "到期时间"])]) >= 2,
        contains_any(task2, ["Name Server", "DNS", "其他重要字段"]),
        contains_any(task2, ["能说明什么", "这类结果能说明什么"]),
        contains_any(task2, ["不能替代", "不能替代什么"]),
        contains_any(task2, ["最小机制说明"]),
    ]
    t2_global_checks = [
        contains_any(normalized, ["查询对象"]),
        sum([contains_any(normalized, ["Registrar", "注册商"]), contains_any(normalized, ["Creation Date", "注册时间"]), contains_any(normalized, ["Expiration Date", "到期时间"])]) >= 2,
        contains_any(normalized, ["Name Server", "DNS", "其他重要字段"]),
        contains_any(normalized, ["能说明什么", "这类结果能说明什么"]),
        contains_any(normalized, ["不能替代", "不能替代什么"]),
        contains_any(normalized, ["最小机制说明"]) and contains_any(normalized, ["Whois", "静态备案", "实时探测"]),
    ]
    t3_checks = [
        contains_any(task3, ["查询对象", "查询条件"]),
        contains_any(task3, ["IP", "主机标识"]) or bool(IP_RE.search(task3)),
        contains_any(task3, ["端口"]) and contains_any(task3, ["服务"]),
        contains_any(task3, ["第三方索引"]),
        contains_any(task3, ["不能替代授权扫描", "不能替代授权", "不能替代扫描"]),
        contains_any(task3, ["最小机制说明"]),
    ]
    t3_global_checks = [
        contains_any(normalized, ["查询对象", "查询条件"]),
        contains_any(normalized, ["IP", "主机标识"]) or bool(IP_RE.search(normalized)),
        contains_any(normalized, ["端口"]) and contains_any(normalized, ["服务"]),
        contains_any(normalized, ["第三方索引"]),
        contains_any(normalized, ["不能替代授权扫描", "不能替代授权", "不能替代扫描"]),
        contains_any(normalized, ["最小机制说明"]) and contains_any(normalized, ["Shodan", "第三方索引结果", "历史索引"]),
    ]
    task4_numbered = count_numbered_items(task4)
    t4_checks = [
        contains_any(task4, ["案例名称", "材料来源", "案例"]),
        contains_any(task4, ["泄露点"]) and (task4_numbered >= 2 or count_hits(task4, ["泄露点"]) >= 2),
        contains_any(task4, ["后果"]),
        contains_any(task4, ["防范建议"]) and task4_numbered >= 3,
        contains_any(task4, ["伦理边界", "不能做利用", "不能做利用与模仿"]),
        contains_any(task4, ["最小机制说明"]) and contains_any(task4, ["判断偏差", "流程缺口", "信任偏差", "权威服从"]),
    ]

    burp_method = contains_any(task5, ["Method", "请求方法"])
    burp_url = contains_any(task5, ["URL", "请求URL", "Request URL"]) or bool(URL_RE.search(task5))
    burp_params = contains_any(task5, ["参数", "Body", "query", "Query String", "username", "password"])
    burp_header = contains_any(task5, ["Header", "Cookie", "Set-Cookie", "User-Agent"])
    burp_status = contains_any(task5, ["状态码", "Status", "302", "200"])
    burp_return = contains_any(task5, ["返回内容", "响应内容", "大致含义", "重定向"])
    burp_evidence_reason = contains_any(task5, ["更接近技术证据", "真实交互", "真实HTTP", "原始数据", "可重现", "可验证"])
    burp_mechanism = contains_any(task5, ["最小机制说明"]) and contains_any(task5, ["代理", "中间人", "监听端口", "转发", "拦截", "请求与响应"])
    t5_checks = [
        contains_any(task5, ["抓包目标"]) and burp_method and burp_url,
        burp_params,
        burp_header,
        burp_status,
        burp_return,
        burp_evidence_reason,
        burp_mechanism,
    ]

    nmap_target = contains_any(task6, ["扫描目标"])
    nmap_command = contains_any(task6, ["扫描命令", "nmap", "namp"])
    nmap_port = contains_any(task6, ["端口"])
    nmap_state = contains_any(task6, ["状态", "open", "closed", "filtered"])
    nmap_service = contains_any(task6, ["服务", "httpd", "nginx", "apache", "microsoft"])
    nmap_active = contains_any(task6, ["主动探测", "主动发起", "探测数据包", "探测报文"])
    nmap_not_vuln = contains_any(task6, ["不等于", "不代表", "已发现漏洞", "开放端口", "服务"])
    nmap_mechanism = contains_any(task6, ["最小机制说明"]) and contains_any(task6, ["SYN", "RST", "响应", "指纹", "推断", "探测包"])
    t6_checks = [
        nmap_target and nmap_command,
        nmap_port and nmap_state and nmap_service,
        nmap_active,
        nmap_not_vuln,
        nmap_mechanism,
    ]

    task_scores = {
        "task1": max_ratio_score(t1_checks, t1_global_checks, profile.tasks[0].score),
        "task2": max_ratio_score(t2_checks, t2_global_checks, profile.tasks[1].score),
        "task3": max_ratio_score(t3_checks, t3_global_checks, profile.tasks[2].score),
        "task4": task_ratio_score(t4_checks, profile.tasks[3].score),
        "task5": task_ratio_score(t5_checks, profile.tasks[4].score),
        "task6": task_ratio_score(t6_checks, profile.tasks[5].score),
    }

    relation_eval = evaluate_graph_section_chapter2(normalized, relation_section, profile.relation_item_score)
    graph_present = relation_eval["status"] == "complete"
    graph_submitted = bool(relation_eval["submitted"])
    self_eval_present = bool(self_eval_section) and contains_any(self_eval_section, ["任务完成情况", "机制理解情况", "证据质量情况", "合规与边界", "自评结论"])
    ai_review_trace = detect_ai_review_trace(ai_review_section or normalized)
    task_completion = clamp(sum(task_scores.values()) + relation_eval["score"] + (profile.self_eval_score if self_eval_present else 0.0), 0.0, 20.0)

    screenshot_markers = count_regex(normalized, SCREENSHOT_MARKER_RE)
    evidence_categories = sum(
        [
            1 if task_scores["task1"] >= 1.5 else 0,
            1 if task_scores["task2"] >= 1.5 else 0,
            1 if task_scores["task3"] >= 1.5 else 0,
            1 if sum([burp_method, burp_url, burp_status, burp_params, burp_header]) >= 3 else 0,
            1 if sum([nmap_target, nmap_command, nmap_port, nmap_state, nmap_service]) >= 3 else 0,
        ]
    )
    artifact_categories = sum(
        [
            1 if bool(URL_RE.search(normalized)) or bool(DOMAIN_RE.search(normalized)) else 0,
            1 if bool(IP_RE.search(normalized)) else 0,
            1 if bool(DATE_RE.search(normalized)) else 0,
            1 if "nmap" in lower or "namp" in lower else 0,
            1 if contains_any(normalized, ["302", "200", "443", "80", "Registrar", "MarkMonitor"]) else 0,
        ]
    )
    evidence_quality = clamp(
        page_points(page_count)
        + screenshot_points(screenshot_markers)
        + ratio_score(evidence_categories, 5, 4.5)
        + ratio_score(artifact_categories, 5, 4.5),
        0.0,
        18.0,
    )

    unified_mechanism_hits = [
        contains_any(tool_mechanism_section, ["Google"]),
        contains_any(tool_mechanism_section, ["Whois"]),
        contains_any(tool_mechanism_section, ["Shodan"]),
        contains_any(tool_mechanism_section, ["社工", "社交工程"]),
        contains_any(tool_mechanism_section, ["Burp"]),
        contains_any(tool_mechanism_section, ["Nmap"]),
    ]
    mechanism_task_hits = max(
        sum(
            [
                contains_any(task1, ["最小机制说明"]),
                contains_any(task2, ["最小机制说明"]),
                contains_any(task3, ["最小机制说明"]),
                contains_any(task4, ["最小机制说明"]),
                contains_any(task5, ["最小机制说明"]),
                contains_any(task6, ["最小机制说明"]),
            ]
        ),
        sum(1 for hit in unified_mechanism_hits if hit),
    )
    mechanism_task_points = ratio_score(mechanism_task_hits, 6, 8.0)
    mechanism_section_present = bool(mechanism_section) and len(mechanism_section) >= 120
    mechanism_section_points = 4.0 if mechanism_section_present else (2.0 if "机制解释" in normalized else 0.0)
    distinction_checks = [
        contains_any(normalized, ["线索", "证据"]),
        contains_any(normalized, ["第三方索引", "索引线索"]),
        contains_any(normalized, ["主动探测", "主动探测结果"]),
        contains_any(normalized, ["Burp", "Nmap"]),
    ]
    distinction_points = ratio_score(sum(distinction_checks), len(distinction_checks), 4.0)
    order_checks = [
        contains_any(normalized, ["被动观察", "被动线索"]),
        contains_any(normalized, ["主动探测后置", "主动探测", "后置"]),
        contains_any(normalized, ["顺序", "优先"]),
        contains_any(normalized, ["更接近证据", "更接近技术证据"]),
    ]
    order_points = ratio_score(sum(order_checks), len(order_checks), 4.0)
    mechanism_score = clamp(mechanism_task_points + mechanism_section_points + distinction_points + order_points, 0.0, 20.0)

    social_root_cause = contains_any(task4, ["判断偏差", "流程缺口", "信任偏差", "权威服从", "路径依赖"])
    risk_section_present = bool(risk_section) and len(risk_section) >= 80
    boundary_present = bool(boundary_section) and len(boundary_section) >= 80
    risk_checks = [
        social_root_cause and contains_any(task4, ["后果"]),
        risk_section_present and contains_any(risk_section, ["风险", "防护", "授权", "误判", "暴露", "边界"]),
        contains_any(task6, ["开放端口", "不等于", "不代表", "漏洞"]),
    ]
    risk_score = ratio_score(sum(risk_checks), len(risk_checks), 12.0)

    suggestion_count = count_numbered_items(task4 + "\n" + risk_section)
    suggestion_points = 5.0 if suggestion_count >= 3 and contains_any(task4, ["防范建议"]) else (3.0 if suggestion_count >= 2 and contains_any(task4, ["防范建议"]) else 0.0)
    control_hits = count_hits(
        normalized,
        ["二次确认", "多级审批", "培训", "脱敏", "HttpOnly", "Secure", "HTTPS", "隐私保护", "限制公网服务暴露", "授权", "最小暴露", "双重验证"],
    )
    controls_points = 5.0 if control_hits >= 3 else (3.0 if control_hits >= 2 else (1.5 if control_hits >= 1 else 0.0))
    protection_score = clamp(suggestion_points + controls_points, 0.0, 10.0)

    ordered_tasks = all(bool(value) for value in [task1, task2, task3, task4, task5, task6])
    section_positions = [
        normalized.find(fragment)
        for fragment in ["Google", "Whois", "Shodan", "社交工程", "Burp", "Nmap"]
        if normalized.find(fragment) != -1
    ]
    order_ok = ordered_tasks or section_positions == sorted(section_positions)
    structure_points = 5.0 if order_ok and graph_present else (4.0 if order_ok and graph_submitted else (3.5 if order_ok else 2.0))
    text_length = len(normalized)
    expression_points = 5.0 if text_length >= 2600 else (4.0 if text_length >= 1800 else (3.0 if text_length >= 1200 else 1.0))
    expression_score = clamp(structure_points + expression_points, 0.0, 10.0)

    professional_review = chapter2_professional_review(
        normalized,
        boundary_section,
        task3,
        task5,
        task6,
        nmap_not_vuln,
        burp_evidence_reason,
        ai_review_trace,
    )
    professional_score = professional_review["score"]

    base_score = task_completion + evidence_quality + mechanism_score + risk_score + protection_score + expression_score + professional_score

    content_issues: list[str] = []
    if sum(1 for value in [task1, task2, task3, task4, task5, task6] if value) < 4:
        content_issues.append("六个任务不完整")
    if relation_eval["status"] == "missing":
        content_issues.append("未提交综合关系图/表")
    elif relation_eval["status"] == "partial":
        content_issues.append("综合关系图已提交但关系表达不足")
    if not self_eval_present:
        content_issues.append("缺少学生自评表")
    if mechanism_task_hits < 3 or not mechanism_section_present:
        content_issues.append("机制解释不完整")
    if not bool(tool_mechanism_section):
        content_issues.append("缺少每个工具的最小机制说明")
    if sum([burp_method, burp_url, burp_status, burp_params, burp_header]) < 3:
        content_issues.append("Burp 关键证据不足")
    if sum([nmap_target, nmap_command, nmap_port, nmap_state, nmap_service]) < 3:
        content_issues.append("Nmap 关键证据不足")
    if not ai_review_trace["present"]:
        content_issues.append("缺少AI输出审核与人工复核记录")
    if professional_score < 7 and professional_review["professional_errors"]:
        content_issues.append("专业校验与责任意识不足")
    if similarity_review and similarity_review.get("is_highly_similar"):
        content_issues.append("与同班作业高度雷同")
    elif similarity_review and similarity_review.get("is_suspected_similar"):
        content_issues.append("存在作业雷同风险，建议人工复核")

    hard_gate_reasons: list[str] = []
    if SESSION_LEAK_RE.search(normalized) and not MASK_HINT_RE.search(normalized):
        hard_gate_reasons.append("敏感信息疑似未脱敏")
    scan_target = find_label_value(task6, ["扫描目标"])
    scan_ip_match = IP_RE.search(scan_target)
    if scan_ip_match and not is_private_target(scan_ip_match.group(0)):
        hard_gate_reasons.append("主动探测或漏洞验证目标授权边界不清")
    if similarity_review and similarity_review.get("is_highly_similar"):
        hard_gate_reasons.append("与同班作业高度雷同")

    semantic_completion_hits = sum(1 for value in task_scores.values() if value >= 1.8)
    high_quality_bundle = (
        graph_present
        and self_eval_present
        and boundary_present
        and mechanism_section_present
        and bool(tool_mechanism_section)
        and semantic_completion_hits >= 6
        and sum([burp_method, burp_url, burp_status, burp_params, burp_header]) >= 4
        and sum([nmap_target, nmap_command, nmap_port, nmap_state, nmap_service]) >= 3
        and professional_score >= 8
    )

    final_score = int(round(clamp(base_score, 0.0, 100.0)))
    if high_quality_bundle and not hard_gate_reasons:
        calibrated_floor = 87
        if evidence_quality >= 15:
            calibrated_floor += 1
        if mechanism_score >= 18:
            calibrated_floor += 1
        if risk_score >= 9 and protection_score >= 8:
            calibrated_floor += 1
        if professional_score >= 9:
            calibrated_floor += 1
        if page_count >= 10:
            calibrated_floor += 1
        if screenshot_markers >= 8 or artifact_categories >= 5:
            calibrated_floor += 1
        final_score = max(final_score, min(calibrated_floor, 93))
    if hard_gate_reasons:
        final_score = min(final_score, 59)

    debug = {
        "rubric_used": "projects5_v5_generic_chapter02",
        "task_completion": round(task_completion, 2),
        "evidence_quality": round(evidence_quality, 2),
        "mechanism_score": round(mechanism_score, 2),
        "risk_score": round(risk_score, 2),
        "protection_score": round(protection_score, 2),
        "expression_score": round(expression_score, 2),
        "professional_score": round(professional_score, 2),
        "base_score": round(base_score, 2),
        "final_score": final_score,
        "page_count": page_count,
        "text_length": text_length,
        "relation_status": relation_eval["status"],
        "relation_header_hits": relation_eval["header_hits"],
        "relation_signal_hits": relation_eval["layer_hits"],
        "relation_relation_hits": relation_eval["relation_hits"],
        "self_eval_present": self_eval_present,
        "ai_review_present": ai_review_trace["present"],
        "ai_review_complete": ai_review_trace["complete"],
        "boundary_present": boundary_present,
        "risk_section_present": risk_section_present,
        "mechanism_section_present": mechanism_section_present,
        "tool_mechanism_section_present": bool(tool_mechanism_section),
        "mechanism_task_hits": mechanism_task_hits,
        "task_scores": {key: round(value, 2) for key, value in task_scores.items()},
        "task_max_scores": {f"task{index}": task.score for index, task in enumerate(profile.tasks, start=1)},
        "task_name_map": {f"task{index}": task.name for index, task in enumerate(profile.tasks, start=1)},
        "task_requirement_map": {f"task{index}": task.semantic_requirements for index, task in enumerate(profile.tasks, start=1)},
        "screenshot_markers": screenshot_markers,
        "evidence_categories": evidence_categories,
        "artifact_categories": artifact_categories,
        "burp_core_hits": sum([burp_method, burp_url, burp_status, burp_params, burp_header]),
        "nmap_core_hits": sum([nmap_target, nmap_command, nmap_port, nmap_state, nmap_service]),
        "semantic_completion_hits": semantic_completion_hits,
        "high_quality_bundle": high_quality_bundle,
        "content_issues": content_issues,
        "hard_gate_reasons": hard_gate_reasons,
        "scan_target": scan_target,
        "domains": extract_domains(normalized)[:8],
        "ai_like_paragraphs": professional_review["ai_like_paragraphs"],
        "unchecked_traces": professional_review["unchecked_traces"],
        "professional_errors": professional_review["professional_errors"],
        "absolute_hits": professional_review["absolute_hits"],
        "similarity_review": similarity_review or {
            "is_suspected_similar": False,
            "is_highly_similar": False,
            "max_full_ratio": 0.0,
            "max_longest_block_ratio": 0.0,
            "matched_peers": [],
        },
    }
    return final_score, debug


def chapter3_scoring(text: str, page_count: int, profile: ChapterProfile, similarity_review: dict[str, Any] | None) -> tuple[int, dict[str, Any]]:
    normalized = normalize_text(text)
    sections = locate_sections_generic(normalized, profile)

    task1 = sections.get("task1", "") or find_context_window(normalized, ["数据库", "查询条件", "返回结果"])
    task2 = sections.get("task2", "") or find_context_window(normalized, ["SQL Injection Low", "DVWA", "现象", "返回差异"])
    task3 = sections.get("task3", "") or find_context_window(normalized, ["拼接", "解析", "执行", "盲注", "Medium"])
    task4 = sections.get("task4", "") or find_context_window(normalized, ["参数化查询", "错误处理", "最小权限", "防护清单"])
    relation_section = sections.get("relation", "") or find_context_window(normalized, ["因果链", "防护关系图", "输入", "拼接", "执行"])
    boundary_section = sections.get("boundary", "") or find_context_window(normalized, ["授权边界", "授权", "允许", "禁止", "脱敏"])
    risk_section = sections.get("risk", "") or find_context_window(normalized, ["风险", "防护", "暴露面", "后果范围"])
    ai_review_section = sections.get("ai_review", "") or find_context_window(normalized, ["AI输出审核", "人工复核", "人工审核痕迹"])
    self_eval_section = sections.get("self_eval", "") or find_context_window(normalized, ["学生自评表", "自评表"])

    t1_checks = [
        contains_any(task1, ["输入条件", "查询条件", "条件"]),
        contains_any(task1, ["数据库"]),
        contains_any(task1, ["查询"]),
        contains_any(task1, ["返回结果", "结果", "返回"]),
        contains_any(task1, ["输入如何进入查询条件", "进入查询条件", "页面功能与数据库有关"]),
    ]
    t1_global = [
        contains_any(normalized, ["输入条件", "查询条件", "条件"]),
        contains_any(normalized, ["数据库"]),
        contains_any(normalized, ["查询"]),
        contains_any(normalized, ["返回结果", "结果", "返回"]),
        contains_any(normalized, ["输入如何进入查询条件", "进入查询条件", "页面功能与数据库有关"]),
    ]

    t2_checks = [
        contains_any(task2, ["DVWA", "SQL Injection Low", "Low"]),
        contains_any(task2, ["普通输入", "正常输入"]),
        contains_any(task2, ["异常输入", "不同输入"]),
        contains_any(task2, ["返回差异", "结果变化", "回显差异"]),
        contains_any(task2, ["不是普通查询", "不像正常查询变化", "改变查询逻辑"]),
    ]
    t2_global = [
        contains_any(normalized, ["DVWA", "SQL Injection Low", "Low"]),
        contains_any(normalized, ["普通输入", "正常输入"]),
        contains_any(normalized, ["异常输入", "不同输入"]),
        contains_any(normalized, ["返回差异", "结果变化", "回显差异"]),
        contains_any(normalized, ["不是普通查询", "不像正常查询变化", "改变查询逻辑"]),
    ]

    t3_checks = [
        contains_any(task3, ["输入"]) and contains_any(task3, ["拼接"]) and contains_any(task3, ["解析"]) and contains_any(task3, ["执行"]),
        contains_any(task3, ["Low"]) and contains_any(task3, ["Medium"]),
        contains_any(task3, ["盲注", "Blind"]),
        contains_any(task3, ["不是字符本身", "结构被污染", "根因", "结构"]),
        contains_any(task3, ["行为差异", "回显", "真假反馈"]),
    ]
    t3_global = [
        contains_any(normalized, ["输入"]) and contains_any(normalized, ["拼接"]) and contains_any(normalized, ["解析"]) and contains_any(normalized, ["执行"]),
        contains_any(normalized, ["Low"]) and contains_any(normalized, ["Medium"]),
        contains_any(normalized, ["盲注", "Blind"]),
        contains_any(normalized, ["不是字符本身", "结构被污染", "根因", "结构"]),
        contains_any(normalized, ["行为差异", "回显", "真假反馈"]),
    ]

    t4_checks = [
        contains_any(task4, ["参数化查询", "Prepared Statement"]),
        contains_any(task4, ["输入校验"]),
        contains_any(task4, ["错误处理"]),
        contains_any(task4, ["最小权限"]),
        contains_any(task4, ["根因", "暴露面", "后果范围"]),
    ]
    t4_global = [
        contains_any(normalized, ["参数化查询", "Prepared Statement"]),
        contains_any(normalized, ["输入校验"]),
        contains_any(normalized, ["错误处理"]),
        contains_any(normalized, ["最小权限"]),
        contains_any(normalized, ["根因", "暴露面", "后果范围"]),
    ]

    task_scores = {
        "task1": max_ratio_score(t1_checks, t1_global, profile.tasks[0].score),
        "task2": max_ratio_score(t2_checks, t2_global, profile.tasks[1].score),
        "task3": max_ratio_score(t3_checks, t3_global, profile.tasks[2].score),
        "task4": max_ratio_score(t4_checks, t4_global, profile.tasks[3].score),
    }

    relation_eval = evaluate_relation_section(
        relation_section,
        ["输入", "拼接", "解析", "执行", "回显", "行为差异", "参数化查询", "最小权限", "防护"],
        profile.relation_item_score,
    )
    self_eval_present = bool(self_eval_section) and contains_any(self_eval_section, ["任务完成情况", "机制理解", "证据质量", "自评结论"])
    ai_review_trace = detect_ai_review_trace(ai_review_section or normalized)
    task_completion = clamp(sum(task_scores.values()) + relation_eval["score"] + (profile.self_eval_score if self_eval_present else 0.0), 0.0, 20.0)

    screenshot_markers = count_regex(normalized, SCREENSHOT_MARKER_RE)
    evidence_categories = sum(
        [
            1 if task_scores["task1"] >= 2.0 else 0,
            1 if task_scores["task2"] >= 2.8 else 0,
            1 if task_scores["task3"] >= 3.2 else 0,
            1 if task_scores["task4"] >= 2.4 else 0,
        ]
    )
    artifact_categories = sum(
        [
            1 if contains_any(normalized, ["DVWA", "SQL", "数据库", "查询"]) else 0,
            1 if contains_any(normalized, ["Low", "Medium", "Blind", "盲注"]) else 0,
            1 if contains_any(normalized, ["回显", "返回差异", "行为差异", "真假反馈"]) else 0,
            1 if contains_any(normalized, ["参数化查询", "错误处理", "最小权限"]) else 0,
            1 if bool(URL_RE.search(normalized)) or bool(IP_RE.search(normalized)) else 0,
        ]
    )
    evidence_quality = clamp(
        page_points(page_count)
        + screenshot_points(screenshot_markers)
        + ratio_score(evidence_categories, 4, 4.5)
        + ratio_score(artifact_categories, 5, 4.5),
        0.0,
        18.0,
    )

    mechanism_checks = [
        contains_any(normalized, ["输入"]) and contains_any(normalized, ["拼接"]) and contains_any(normalized, ["解析"]) and contains_any(normalized, ["执行"]),
        contains_any(normalized, ["结构被污染", "根因", "输入进入了不该进入的结构"]),
        contains_any(normalized, ["Low"]) and contains_any(normalized, ["Medium"]),
        contains_any(normalized, ["盲注", "Blind"]) and contains_any(normalized, ["行为差异", "真假反馈", "无直接回显"]),
        contains_any(normalized, ["参数化查询", "最小权限", "错误处理"]),
    ]
    mechanism_score = ratio_score(sum(mechanism_checks), len(mechanism_checks), 20.0)

    boundary_present = bool(boundary_section) and contains_any(boundary_section, ["授权", "允许", "禁止", "脱敏"])
    risk_checks = [
        boundary_present,
        contains_any(task3 + "\n" + normalized, ["根因", "结构被污染", "不是字符本身"]),
        contains_any(task4 + "\n" + risk_section, ["根因", "暴露面", "后果范围"]),
    ]
    risk_score = ratio_score(sum(risk_checks), len(risk_checks), 12.0)

    control_hits = count_hits(normalized, ["参数化查询", "输入校验", "错误处理", "最小权限", "结构分离", "防护清单"])
    protection_score = 10.0 if control_hits >= 5 else (8.0 if control_hits >= 4 else (6.0 if control_hits >= 3 else (3.0 if control_hits >= 2 else 0.0)))

    ordered_sections = all(bool(sections.get(f"task{index}")) for index in range(1, len(profile.tasks) + 1))
    structure_points = 5.0 if ordered_sections and relation_eval["status"] == "complete" else (4.0 if ordered_sections else 2.5)
    text_length = len(normalized)
    expression_points = 5.0 if text_length >= 2400 else (4.0 if text_length >= 1700 else (3.0 if text_length >= 1100 else 1.5))
    expression_score = clamp(structure_points + expression_points, 0.0, 10.0)

    public_targets: list[str] = []
    for candidate in extract_domains(task2 + "\n" + task3):
        if not is_private_target(candidate) and "dvwa" not in candidate and "localhost" not in candidate:
            public_targets.append(candidate)
    for ip in IP_RE.findall(task2 + "\n" + task3):
        if not is_private_target(ip):
            public_targets.append(ip)

    professional_review = build_professional_review(
        normalized,
        ai_review_trace,
        checks=[
            {
                "ok": contains_any(task2 + "\n" + task3 + "\n" + normalized, ["改变查询逻辑", "结构被污染", "不是普通查询变化"]),
                "item": "未区分正常查询变化与查询逻辑被改变",
                "why": "第三章核心不是“结果变了”，而是判断结果变化是否意味着查询结构被输入影响。",
                "fix": "补写：关键不在结果不同，而在输入影响了查询逻辑或数据库解释方式。",
                "penalty": 2.0,
            },
            {
                "ok": not contains_any(normalized, ["payload本身就是根因", "特殊字符本身导致漏洞"]) and contains_any(normalized, ["根因", "结构", "不是字符本身", "结构被污染"]),
                "item": "把 payload 或特殊字符误写为根因",
                "why": "第三章要建立“字符只是现象入口，根因是数据进入结构”的认知。",
                "fix": "改写为：payload 只是触发现象，真正根因是输入进入了不该进入的 SQL 结构。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(normalized, ["盲注", "Blind"]) and contains_any(normalized, ["同一根因", "证据形式不同", "行为差异", "无直接回显"]) or not contains_any(normalized, ["盲注", "Blind"]),
                "item": "盲注与普通 SQL 注入关系说明不足",
                "why": "第三章不能把盲注当成完全不同的问题，而应理解为同一根因在证据层变弱后的表现。",
                "fix": "补写：盲注与普通 SQL 注入共享同一根因，只是观察证据从直接回显变成行为差异或真假反馈。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(task4 + "\n" + normalized, ["参数化查询", "输入校验", "错误处理", "最小权限", "根因", "暴露面", "后果范围"]),
                "item": "防护层次区分不足",
                "why": "第三章需要把参数化、过滤、错误处理、最小权限区分到不同控制层，而不是平铺罗列。",
                "fix": "补写：参数化查询控制根因，错误处理控制暴露面，最小权限控制后果范围。",
                "penalty": 2.0,
            },
            {
                "ok": boundary_present,
                "item": "授权边界与风险表达不足",
                "why": "SQL 注入实验必须限定在 DVWA 或教师授权环境，否则会直接越过教学边界。",
                "fix": "补写：实验仅限 DVWA、本机或教师授权环境，禁止对真实未授权站点做验证。",
                "penalty": 2.0,
            },
        ],
    )
    professional_score = professional_review["score"]

    base_score = task_completion + evidence_quality + mechanism_score + risk_score + protection_score + expression_score + professional_score

    content_issues: list[str] = []
    for key, value in task_scores.items():
        max_score = next(task.score for index, task in enumerate(profile.tasks, start=1) if f"task{index}" == key)
        if value < 0.6 * max_score:
            content_issues.append(f"{profile.tasks[int(key[-1]) - 1].name}内容不全")
    if relation_eval["status"] == "missing":
        content_issues.append(f"未提交{profile.relation_item_name}")
    elif relation_eval["status"] == "partial":
        content_issues.append(f"{profile.relation_item_name}已提交但关系表达不足")
    if not self_eval_present:
        content_issues.append("缺少学生自评表")
    if not ai_review_trace["present"]:
        content_issues.append("缺少AI输出审核与人工复核记录")
    if professional_score < 7 and professional_review["professional_errors"]:
        content_issues.append("专业校验与责任意识不足")
    if similarity_review and similarity_review.get("is_highly_similar"):
        content_issues.append("与同班作业高度雷同")
    elif similarity_review and similarity_review.get("is_suspected_similar"):
        content_issues.append("存在作业雷同风险，建议人工复核")

    hard_gate_reasons: list[str] = []
    if SESSION_LEAK_RE.search(normalized) and not MASK_HINT_RE.search(normalized):
        hard_gate_reasons.append("敏感信息疑似未脱敏")
    if public_targets:
        hard_gate_reasons.append("主动探测或漏洞验证目标授权边界不清")
    if similarity_review and similarity_review.get("is_highly_similar"):
        hard_gate_reasons.append("与同班作业高度雷同")

    final_score = int(round(clamp(base_score, 0.0, 100.0)))
    if hard_gate_reasons:
        final_score = min(final_score, 59)

    debug = {
        "rubric_used": "projects5_v5_generic_chapter03",
        "task_completion": round(task_completion, 2),
        "evidence_quality": round(evidence_quality, 2),
        "mechanism_score": round(mechanism_score, 2),
        "risk_score": round(risk_score, 2),
        "protection_score": round(protection_score, 2),
        "expression_score": round(expression_score, 2),
        "professional_score": round(professional_score, 2),
        "base_score": round(base_score, 2),
        "final_score": final_score,
        "page_count": page_count,
        "text_length": text_length,
        "relation_status": relation_eval["status"],
        "relation_signal_hits": relation_eval["signal_hits"],
        "relation_relation_hits": relation_eval["relation_hits"],
        "self_eval_present": self_eval_present,
        "ai_review_present": ai_review_trace["present"],
        "ai_review_complete": ai_review_trace["complete"],
        "task_scores": {key: round(value, 2) for key, value in task_scores.items()},
        "task_max_scores": {f"task{index}": task.score for index, task in enumerate(profile.tasks, start=1)},
        "task_name_map": {f"task{index}": task.name for index, task in enumerate(profile.tasks, start=1)},
        "task_requirement_map": {f"task{index}": task.semantic_requirements for index, task in enumerate(profile.tasks, start=1)},
        "screenshot_markers": screenshot_markers,
        "content_issues": content_issues,
        "hard_gate_reasons": hard_gate_reasons,
        "public_targets": public_targets[:5],
        "domains": extract_domains(normalized)[:8],
        "ai_like_paragraphs": professional_review["ai_like_paragraphs"],
        "unchecked_traces": professional_review["unchecked_traces"],
        "professional_errors": professional_review["professional_errors"],
        "absolute_hits": professional_review["absolute_hits"],
        "similarity_review": similarity_review or {
            "is_suspected_similar": False,
            "is_highly_similar": False,
            "max_full_ratio": 0.0,
            "max_longest_block_ratio": 0.0,
            "matched_peers": [],
        },
    }
    return final_score, debug


def chapter4_scoring(text: str, page_count: int, profile: ChapterProfile, similarity_review: dict[str, Any] | None) -> tuple[int, dict[str, Any]]:
    normalized = normalize_text(text)
    sections = locate_sections_generic(normalized, profile)

    task1 = sections.get("task1", "") or find_context_window(normalized, ["浏览器", "解析", "输出位置", "上下文", "XSS"])
    task2 = sections.get("task2", "") or find_context_window(normalized, ["Reflected XSS", "反射型", "Low", "即时回显"])
    task3 = sections.get("task3", "") or find_context_window(normalized, ["Stored XSS", "存储型", "Low", "保存"])
    task4 = sections.get("task4", "") or find_context_window(normalized, ["DOM XSS", "DOM", "前端", "类型差异"])
    task5 = sections.get("task5", "") or find_context_window(normalized, ["Medium", "输出编码", "CSP", "防护清单"])
    relation_section = sections.get("relation", "") or find_context_window(normalized, ["关系图", "关系表", "风险链", "输入", "输出", "解析", "执行"])
    boundary_section = sections.get("boundary", "") or find_context_window(normalized, ["授权边界", "授权", "允许", "禁止", "脱敏"])
    risk_section = sections.get("risk", "") or find_context_window(normalized, ["风险", "传播", "防护", "控制点"])
    ai_review_section = sections.get("ai_review", "") or find_context_window(normalized, ["AI输出审核", "人工复核", "人工审核痕迹"])
    self_eval_section = sections.get("self_eval", "") or find_context_window(normalized, ["学生自评表", "自评表"])

    t1_checks = [
        contains_any(task1, ["浏览器"]) and contains_any(task1, ["解析", "解释"]),
        contains_any(task1, ["输出位置", "上下文", "文本区", "属性区", "脚本区", "DOM"]),
        contains_any(task1, ["同样", "不同位置", "不同结果", "不同风险"]),
        contains_any(task1, ["风险", "安全关注点", "控制点"]),
    ]
    t1_global = [
        contains_any(normalized, ["浏览器"]) and contains_any(normalized, ["解析", "解释"]),
        contains_any(normalized, ["输出位置", "上下文", "文本区", "属性区", "脚本区", "DOM"]),
        contains_any(normalized, ["同样", "不同位置", "不同结果", "不同风险"]),
        contains_any(normalized, ["风险", "安全关注点", "控制点"]),
    ]

    t2_checks = [
        contains_any(task2, ["Reflected XSS", "反射型", "Reflected"]),
        contains_any(task2, ["Low"]),
        contains_any(task2, ["输入点", "即时回显", "回显", "提交后立即"]),
        contains_any(task2, ["执行", "页面变化", "浏览器执行", "危险位置"]),
    ]
    t2_global = [
        contains_any(normalized, ["Reflected XSS", "反射型", "Reflected"]),
        contains_any(normalized, ["Low"]),
        contains_any(normalized, ["输入点", "即时回显", "回显", "提交后立即"]),
        contains_any(normalized, ["执行", "页面变化", "浏览器执行", "危险位置"]),
    ]

    t3_checks = [
        contains_any(task3, ["Stored XSS", "存储型", "Stored"]),
        contains_any(task3, ["Low"]),
        contains_any(task3, ["保存", "存储", "后续访问", "持续"]),
        contains_any(task3, ["与反射型不同", "影响更持续", "传播", "持续影响"]),
    ]
    t3_global = [
        contains_any(normalized, ["Stored XSS", "存储型", "Stored"]),
        contains_any(normalized, ["Low"]),
        contains_any(normalized, ["保存", "存储", "后续访问", "持续"]),
        contains_any(normalized, ["与反射型不同", "影响更持续", "传播", "持续影响"]),
    ]

    t4_checks = [
        contains_any(task4, ["DOM XSS", "DOM"]),
        contains_any(task4, ["前端", "DOM更新", "浏览器端", "客户端"]),
        contains_any(task4, ["类型差异", "与 Reflected", "与 Stored", "差异"]),
        contains_any(task4, ["输入"]) and contains_any(task4, ["输出", "解析"]),
    ]
    t4_global = [
        contains_any(normalized, ["DOM XSS", "DOM"]),
        contains_any(normalized, ["前端", "DOM更新", "浏览器端", "客户端"]),
        contains_any(normalized, ["类型差异", "与 Reflected", "与 Stored", "差异"]),
        contains_any(normalized, ["输入"]) and contains_any(normalized, ["输出", "解析"]),
    ]

    t5_checks = [
        contains_any(task5, ["Low"]) and contains_any(task5, ["Medium"]),
        contains_any(task5, ["有限过滤", "过滤", "黑名单", "标签"]),
        contains_any(task5, ["输出编码", "编码", "转义", "上下文安全"]),
        contains_any(task5, ["CSP", "防护清单", "防护建议"]),
    ]
    t5_global = [
        contains_any(normalized, ["Low"]) and contains_any(normalized, ["Medium"]),
        contains_any(normalized, ["有限过滤", "过滤", "黑名单", "标签"]),
        contains_any(normalized, ["输出编码", "编码", "转义", "上下文安全"]),
        contains_any(normalized, ["CSP", "防护清单", "防护建议"]),
    ]

    task_scores = {
        "task1": max_ratio_score(t1_checks, t1_global, profile.tasks[0].score),
        "task2": max_ratio_score(t2_checks, t2_global, profile.tasks[1].score),
        "task3": max_ratio_score(t3_checks, t3_global, profile.tasks[2].score),
        "task4": max_ratio_score(t4_checks, t4_global, profile.tasks[3].score),
        "task5": max_ratio_score(t5_checks, t5_global, profile.tasks[4].score),
    }

    relation_eval = evaluate_relation_section(
        relation_section,
        ["输入", "输出", "解析", "执行", "浏览器", "Reflected", "Stored", "DOM", "输出编码", "CSP"],
        profile.relation_item_score,
    )
    self_eval_present = bool(self_eval_section) and contains_any(self_eval_section, ["任务完成情况", "机制理解", "证据质量", "自评结论"])
    ai_review_trace = detect_ai_review_trace(ai_review_section or normalized)

    screenshot_markers = count_regex(normalized, SCREENSHOT_MARKER_RE)
    evidence_categories = sum(
        [
            1 if task_scores["task1"] >= 2.2 else 0,
            1 if task_scores["task2"] >= 2.2 else 0,
            1 if task_scores["task3"] >= 2.2 else 0,
            1 if task_scores["task4"] >= 2.2 else 0,
            1 if task_scores["task5"] >= 2.2 else 0,
        ]
    )
    artifact_categories = sum(
        [
            1 if contains_any(normalized, ["XSS", "浏览器", "上下文", "解析"]) else 0,
            1 if contains_any(normalized, ["Reflected", "反射型", "Stored", "存储型", "DOM XSS"]) else 0,
            1 if contains_any(normalized, ["Low", "Medium"]) else 0,
            1 if contains_any(normalized, ["输出编码", "CSP", "转义", "过滤"]) else 0,
            1 if bool(URL_RE.search(normalized)) or bool(IP_RE.search(normalized)) else 0,
        ]
    )
    evidence_quality = clamp(
        page_points(page_count)
        + screenshot_points(screenshot_markers)
        + ratio_score(evidence_categories, 5, 4.5)
        + ratio_score(artifact_categories, 5, 4.5),
        0.0,
        18.0,
    )

    mechanism_checks = [
        contains_any(normalized, ["浏览器"]) and contains_any(normalized, ["解析", "解释"]) and not contains_any(normalized, ["只是显示", "仅仅显示"]),
        contains_any(normalized, ["输入"]) and contains_any(normalized, ["输出"]) and contains_any(normalized, ["解析"]) and contains_any(normalized, ["执行"]),
        contains_any(normalized, ["Reflected", "反射型"]) and contains_any(normalized, ["Stored", "存储型"]) and contains_any(normalized, ["DOM XSS", "DOM"]),
        contains_any(normalized, ["Low"]) and contains_any(normalized, ["Medium"]) and contains_any(normalized, ["过滤", "上下文", "安全"]),
        contains_any(normalized, ["输出编码", "CSP", "防护清单", "转义"]),
    ]
    mechanism_score = ratio_score(sum(mechanism_checks), len(mechanism_checks), 20.0)

    boundary_present = bool(boundary_section) and contains_any(boundary_section, ["授权", "允许", "禁止", "脱敏"])
    risk_checks = [
        boundary_present,
        contains_any(task3 + "\n" + task4 + "\n" + normalized, ["持续", "传播", "后续访问", "前端环境", "执行"]),
        contains_any(task5 + "\n" + risk_section, ["输出编码", "CSP", "防护", "控制"]),
    ]
    risk_score = ratio_score(sum(risk_checks), len(risk_checks), 12.0)

    control_hits = count_hits(normalized, ["输出编码", "CSP", "转义", "上下文", "过滤", "防护", "编码"])
    protection_score = 10.0 if control_hits >= 6 else (8.0 if control_hits >= 5 else (6.0 if control_hits >= 3 else (3.0 if control_hits >= 2 else 0.0)))

    ordered_sections = all(bool(sections.get(f"task{index}")) for index in range(1, len(profile.tasks) + 1))
    structure_points = 5.0 if ordered_sections and relation_eval["status"] == "complete" else (4.0 if ordered_sections else 2.5)
    text_length = len(normalized)
    expression_points = 5.0 if text_length >= 2600 else (4.0 if text_length >= 1800 else (3.0 if text_length >= 1200 else 1.5))
    expression_score = clamp(structure_points + expression_points, 0.0, 10.0)

    public_targets = detect_public_targets(task2 + "\n" + task3 + "\n" + task4)
    professional_review = build_professional_review(
        normalized,
        ai_review_trace,
        checks=[
            {
                "ok": contains_any(normalized, ["浏览器执行", "浏览器会执行", "浏览器解释"]) and not contains_any(normalized, ["服务器执行脚本"]),
                "item": "未区分浏览器执行与服务器返回",
                "why": "第四章的核心是浏览器如何解释输出，而不是服务器在执行脚本。",
                "fix": "补写：服务器返回内容后，真正解释和执行发生在浏览器端的危险上下文中。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(normalized, ["输出上下文", "上下文失控", "输出位置"]) and not contains_any(normalized, ["只要输入脚本就会xss"]),
                "item": "把 XSS 根因简化成“输入了脚本”",
                "why": "第四章应建立“根因在输出上下文失控”的认知，而不是停留在样例层。",
                "fix": "改写为：XSS 的关键不是输入了什么，而是浏览器把不可信内容放进了会解释执行的位置。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(normalized, ["Reflected", "反射型"]) and contains_any(normalized, ["Stored", "存储型"]) and contains_any(normalized, ["DOM XSS", "DOM"]) and contains_any(normalized, ["差异", "路径"]),
                "item": "三类 XSS 的路径差异说明不足",
                "why": "本章不仅要会做现象记录，还要能比较不同类型的形成路径。",
                "fix": "补写：反射型强调即时回显，存储型强调内容被保存后持续输出，DOM XSS 更接近前端环境中的 DOM 更新路径。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(task5 + "\n" + normalized, ["有限过滤", "输出编码", "CSP", "上下文安全"]),
                "item": "防护层次区分不足",
                "why": "第四章需要区分有限过滤、上下文相关输出编码与 CSP 的不同作用层级。",
                "fix": "补写：过滤主要限制部分输入表面，输出编码控制浏览器解释方式，CSP 属于额外的执行限制层。",
                "penalty": 2.0,
            },
            {
                "ok": boundary_present,
                "item": "授权边界与风险表达不足",
                "why": "XSS 实验和观察必须限定在 DVWA 或教师授权环境，避免把课堂现象迁移到未授权真实目标。",
                "fix": "补写：实验仅限 DVWA、本机或教师授权环境，并说明截图已按要求脱敏。",
                "penalty": 2.0,
            },
        ],
    )

    extra_hard_gate_reasons = ["主动测试目标授权边界不清"] if public_targets else []
    return finalize_specialized_scoring(
        normalized=normalized,
        page_count=page_count,
        profile=profile,
        sections=sections,
        task_scores=task_scores,
        relation_eval=relation_eval,
        self_eval_present=self_eval_present,
        ai_review_trace=ai_review_trace,
        evidence_quality=evidence_quality,
        mechanism_score=mechanism_score,
        risk_score=risk_score,
        protection_score=protection_score,
        expression_score=expression_score,
        professional_review=professional_review,
        similarity_review=similarity_review,
        rubric_used="projects5_v5_generic_chapter04",
        extra_hard_gate_reasons=extra_hard_gate_reasons,
        extra_debug={
            "screenshot_markers": screenshot_markers,
            "relation_signal_hits": relation_eval["signal_hits"],
            "relation_relation_hits": relation_eval["relation_hits"],
            "public_targets": public_targets[:5],
            "domains": extract_domains(normalized)[:8],
        },
    )


def chapter5_scoring(text: str, page_count: int, profile: ChapterProfile, similarity_review: dict[str, Any] | None) -> tuple[int, dict[str, Any]]:
    normalized = normalize_text(text)
    sections = locate_sections_generic(normalized, profile)

    task1 = sections.get("task1", "") or find_context_window(normalized, ["上传", "验证", "存储", "访问", "执行"])
    task2 = sections.get("task2", "") or find_context_window(normalized, ["File Upload", "Low", "上传结果", "保存位置"])
    task3 = sections.get("task3", "") or find_context_window(normalized, ["Medium", "扩展名", "MIME", "差异"])
    task4 = sections.get("task4", "") or find_context_window(normalized, ["防护清单", "重命名", "隔离存储", "禁执行", "访问控制"])
    relation_section = sections.get("relation", "") or find_context_window(normalized, ["链路图", "关系图", "上传", "验证", "存储", "访问", "执行"])
    boundary_section = sections.get("boundary", "") or find_context_window(normalized, ["授权边界", "授权", "允许", "禁止", "脱敏"])
    risk_section = sections.get("risk", "") or find_context_window(normalized, ["风险", "控制点", "防护", "残余风险"])
    ai_review_section = sections.get("ai_review", "") or find_context_window(normalized, ["AI输出审核", "人工复核", "人工审核痕迹"])
    self_eval_section = sections.get("self_eval", "") or find_context_window(normalized, ["学生自评表", "自评表"])

    t1_checks = [
        contains_any(task1, ["上传"]) and contains_any(task1, ["验证"]) and contains_any(task1, ["存储"]) and contains_any(task1, ["访问"]) and contains_any(task1, ["执行"]),
        contains_any(task1, ["风险点", "控制点"]),
        count_hits(task1, ["验证", "存储", "访问", "执行", "控制点", "风险点"]) >= 4,
        contains_any(task1, ["不是一个动作", "是一条链路", "流程控制"]),
    ]
    t1_global = [
        contains_any(normalized, ["上传"]) and contains_any(normalized, ["验证"]) and contains_any(normalized, ["存储"]) and contains_any(normalized, ["访问"]) and contains_any(normalized, ["执行"]),
        contains_any(normalized, ["风险点", "控制点"]),
        count_hits(normalized, ["验证", "存储", "访问", "执行", "控制点", "风险点"]) >= 4,
        contains_any(normalized, ["不是一个动作", "是一条链路", "流程控制"]),
    ]

    t2_checks = [
        contains_any(task2, ["File Upload", "Low"]),
        contains_any(task2, ["上传结果", "上传成功", "保存位置", "保存证据"]),
        contains_any(task2, ["访问路径", "可访问", "访问"]),
        contains_any(task2, ["缺失", "几乎没有控制", "错误信任", "控制点"]),
    ]
    t2_global = [
        contains_any(normalized, ["File Upload", "Low"]),
        contains_any(normalized, ["上传结果", "上传成功", "保存位置", "保存证据"]),
        contains_any(normalized, ["访问路径", "可访问", "访问"]),
        contains_any(normalized, ["缺失", "几乎没有控制", "错误信任", "控制点"]),
    ]

    t3_checks = [
        contains_any(task3, ["Low"]) and contains_any(task3, ["Medium"]),
        contains_any(task3, ["扩展名", "MIME", "内容检查"]),
        contains_any(task3, ["控制了哪一环", "验证", "存储", "执行", "访问"]),
        contains_any(task3, ["单点检查不等于完整防护", "残余风险", "仍然不够"]),
    ]
    t3_global = [
        contains_any(normalized, ["Low"]) and contains_any(normalized, ["Medium"]),
        contains_any(normalized, ["扩展名", "MIME", "内容检查"]),
        contains_any(normalized, ["控制了哪一环", "验证", "存储", "执行", "访问"]),
        contains_any(normalized, ["单点检查不等于完整防护", "残余风险", "仍然不够"]),
    ]

    t4_checks = [
        contains_any(task4, ["扩展名", "MIME", "内容检查"]),
        contains_any(task4, ["重命名", "隔离存储", "禁执行", "访问控制"]),
        count_hits(task4, ["防护", "建议", "控制"]) >= 4 or count_numbered_items(task4) >= 4,
        contains_any(task4, ["对应哪一环", "验证", "存储", "访问", "执行"]),
    ]
    t4_global = [
        contains_any(normalized, ["扩展名", "MIME", "内容检查"]),
        contains_any(normalized, ["重命名", "隔离存储", "禁执行", "访问控制"]),
        count_hits(normalized, ["防护", "建议", "控制"]) >= 4 or count_numbered_items(normalized) >= 4,
        contains_any(normalized, ["对应哪一环", "验证", "存储", "访问", "执行"]),
    ]

    task_scores = {
        "task1": max_ratio_score(t1_checks, t1_global, profile.tasks[0].score),
        "task2": max_ratio_score(t2_checks, t2_global, profile.tasks[1].score),
        "task3": max_ratio_score(t3_checks, t3_global, profile.tasks[2].score),
        "task4": max_ratio_score(t4_checks, t4_global, profile.tasks[3].score),
    }

    relation_eval = evaluate_relation_section(
        relation_section,
        ["上传", "验证", "存储", "访问", "执行", "控制点", "风险点", "隔离存储", "禁执行"],
        profile.relation_item_score,
    )
    self_eval_present = bool(self_eval_section) and contains_any(self_eval_section, ["任务完成情况", "机制理解", "证据质量", "自评结论"])
    ai_review_trace = detect_ai_review_trace(ai_review_section or normalized)

    screenshot_markers = count_regex(normalized, SCREENSHOT_MARKER_RE)
    evidence_categories = sum(
        [
            1 if task_scores["task1"] >= 2.2 else 0,
            1 if task_scores["task2"] >= 2.8 else 0,
            1 if task_scores["task3"] >= 2.8 else 0,
            1 if task_scores["task4"] >= 2.8 else 0,
        ]
    )
    artifact_categories = sum(
        [
            1 if contains_any(normalized, ["File Upload", "上传", "保存位置", "访问路径"]) else 0,
            1 if contains_any(normalized, ["Low", "Medium"]) else 0,
            1 if contains_any(normalized, ["扩展名", "MIME", "内容检查"]) else 0,
            1 if contains_any(normalized, ["重命名", "隔离存储", "禁执行", "访问控制"]) else 0,
            1 if bool(URL_RE.search(normalized)) or bool(IP_RE.search(normalized)) else 0,
        ]
    )
    evidence_quality = clamp(
        page_points(page_count)
        + screenshot_points(screenshot_markers)
        + ratio_score(evidence_categories, 4, 4.5)
        + ratio_score(artifact_categories, 5, 4.5),
        0.0,
        18.0,
    )

    mechanism_checks = [
        contains_any(normalized, ["上传"]) and contains_any(normalized, ["验证"]) and contains_any(normalized, ["存储"]) and contains_any(normalized, ["访问"]) and contains_any(normalized, ["执行"]),
        contains_any(normalized, ["不是一个动作", "是一条链路", "流程控制"]),
        contains_any(normalized, ["Low"]) and contains_any(normalized, ["Medium"]) and contains_any(normalized, ["单点检查不等于完整防护", "控制链"]),
        contains_any(normalized, ["前端"]) and contains_any(normalized, ["服务端"]) or contains_any(normalized, ["服务端控制", "前端限制不能替代"]),
        contains_any(normalized, ["重命名", "隔离存储", "禁执行", "访问控制"]),
    ]
    mechanism_score = ratio_score(sum(mechanism_checks), len(mechanism_checks), 20.0)

    boundary_present = bool(boundary_section) and contains_any(boundary_section, ["授权", "允许", "禁止", "脱敏"])
    risk_checks = [
        boundary_present,
        contains_any(task1 + "\n" + task3 + "\n" + normalized, ["风险点", "控制点", "残余风险", "链路"]),
        contains_any(task4 + "\n" + risk_section, ["隔离存储", "禁执行", "访问控制", "服务端"]),
    ]
    risk_score = ratio_score(sum(risk_checks), len(risk_checks), 12.0)

    control_hits = count_hits(normalized, ["扩展名", "MIME", "内容检查", "重命名", "隔离存储", "禁执行", "访问控制", "服务端"])
    protection_score = 10.0 if control_hits >= 7 else (8.0 if control_hits >= 5 else (6.0 if control_hits >= 4 else (3.0 if control_hits >= 2 else 0.0)))

    ordered_sections = all(bool(sections.get(f"task{index}")) for index in range(1, len(profile.tasks) + 1))
    structure_points = 5.0 if ordered_sections and relation_eval["status"] == "complete" else (4.0 if ordered_sections else 2.5)
    text_length = len(normalized)
    expression_points = 5.0 if text_length >= 2500 else (4.0 if text_length >= 1700 else (3.0 if text_length >= 1100 else 1.5))
    expression_score = clamp(structure_points + expression_points, 0.0, 10.0)

    public_targets = detect_public_targets(task2 + "\n" + task3)
    professional_review = build_professional_review(
        normalized,
        ai_review_trace,
        checks=[
            {
                "ok": contains_any(normalized, ["是一条链路", "流程控制", "上传链路"]),
                "item": "未把文件上传理解为链路问题",
                "why": "第五章要建立“上传—验证—存储—访问—执行”链条，而不是把风险只看成上传成功这一瞬间。",
                "fix": "补写：上传风险来自整条处理链，不只在文件被接收时，还在后续如何保存、访问和执行。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(normalized, ["扩展名", "MIME", "内容检查", "重命名", "隔离存储", "禁执行"]),
                "item": "未区分不同控制点的作用层次",
                "why": "第五章要看控制链是否完整，而不是简单罗列几个检查动作。",
                "fix": "补写：扩展名、MIME、内容检查控制验证环节，重命名与隔离存储控制保存环节，禁执行与访问控制控制后续风险。",
                "penalty": 2.0,
            },
            {
                "ok": not contains_any(normalized, ["上传成功就等于漏洞", "只要上传成功就证明系统不安全"]) and contains_any(normalized, ["上传成功", "后续处理", "访问", "执行"]),
                "item": "把“上传成功”直接写成漏洞结论",
                "why": "上传成功只是现象起点，是否形成高风险取决于后续如何存储、访问和执行。",
                "fix": "改写为：上传成功说明系统接收了用户内容，仍需结合存储位置、访问方式和执行条件判断风险等级。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(normalized, ["前端限制不能替代", "服务端控制", "服务端"]) and contains_any(normalized, ["单点检查不等于完整防护", "控制链"]),
                "item": "未说明前端限制与服务端控制边界",
                "why": "第五章要建立“真正可靠的上传安全依赖服务端完整控制链”的认知。",
                "fix": "补写：前端限制只能提升用户体验，真正安全判断必须落在服务端的验证、存储和执行控制链上。",
                "penalty": 2.0,
            },
            {
                "ok": boundary_present,
                "item": "授权边界与风险表达不足",
                "why": "文件上传实验必须限定在 DVWA 或教师授权环境，避免把课堂观察迁移到未授权真实目标。",
                "fix": "补写：实验仅限 DVWA、本机或教师授权环境，并说明截图与路径已做脱敏处理。",
                "penalty": 2.0,
            },
        ],
    )

    extra_hard_gate_reasons = ["上传测试目标授权边界不清"] if public_targets else []
    return finalize_specialized_scoring(
        normalized=normalized,
        page_count=page_count,
        profile=profile,
        sections=sections,
        task_scores=task_scores,
        relation_eval=relation_eval,
        self_eval_present=self_eval_present,
        ai_review_trace=ai_review_trace,
        evidence_quality=evidence_quality,
        mechanism_score=mechanism_score,
        risk_score=risk_score,
        protection_score=protection_score,
        expression_score=expression_score,
        professional_review=professional_review,
        similarity_review=similarity_review,
        rubric_used="projects5_v5_generic_chapter05",
        extra_hard_gate_reasons=extra_hard_gate_reasons,
        extra_debug={
            "screenshot_markers": screenshot_markers,
            "relation_signal_hits": relation_eval["signal_hits"],
            "relation_relation_hits": relation_eval["relation_hits"],
            "public_targets": public_targets[:5],
            "domains": extract_domains(normalized)[:8],
        },
    )


def chapter6_scoring(text: str, page_count: int, profile: ChapterProfile, similarity_review: dict[str, Any] | None) -> tuple[int, dict[str, Any]]:
    normalized = normalize_text(text)
    sections = locate_sections_generic(normalized, profile)

    task1 = sections.get("task1", "") or find_context_window(normalized, ["应用层", "系统层", "命令执行", "跨层"])
    task2 = sections.get("task2", "") or find_context_window(normalized, ["Command Injection", "Low", "结果变化", "系统层"])
    task3 = sections.get("task3", "") or find_context_window(normalized, ["Medium", "结构", "SQL 注入", "黑名单"])
    task4 = sections.get("task4", "") or find_context_window(normalized, ["参数分离", "最小权限", "安全接口", "改进方案"])
    relation_section = sections.get("relation", "") or find_context_window(normalized, ["风险链路图", "关系图", "应用层", "系统层", "命令", "执行"])
    boundary_section = sections.get("boundary", "") or find_context_window(normalized, ["授权边界", "授权", "允许", "禁止", "脱敏"])
    risk_section = sections.get("risk", "") or find_context_window(normalized, ["风险", "后果范围", "防护", "控制点"])
    ai_review_section = sections.get("ai_review", "") or find_context_window(normalized, ["AI输出审核", "人工复核", "人工审核痕迹"])
    self_eval_section = sections.get("self_eval", "") or find_context_window(normalized, ["学生自评表", "自评表"])

    t1_checks = [
        contains_any(task1, ["应用层"]) and contains_any(task1, ["系统层"]),
        contains_any(task1, ["输入如何进入", "进入系统调用链", "后台调用"]),
        contains_any(task1, ["跨层风险", "边界被放大", "放大"]),
        contains_any(task1, ["业务场景", "探测", "调用系统命令"]),
    ]
    t1_global = [
        contains_any(normalized, ["应用层"]) and contains_any(normalized, ["系统层"]),
        contains_any(normalized, ["输入如何进入", "进入系统调用链", "后台调用"]),
        contains_any(normalized, ["跨层风险", "边界被放大", "放大"]),
        contains_any(normalized, ["业务场景", "探测", "调用系统命令"]),
    ]

    t2_checks = [
        contains_any(task2, ["Command Injection", "Low"]),
        contains_any(task2, ["输入", "结果差异", "返回差异", "系统层"]),
        contains_any(task2, ["不是页面显示不同", "不只是页面回显", "系统执行"]),
        contains_any(task2, ["DVWA", "实验页面", "目标"]),
    ]
    t2_global = [
        contains_any(normalized, ["Command Injection", "Low"]),
        contains_any(normalized, ["输入", "结果差异", "返回差异", "系统层"]),
        contains_any(normalized, ["不是页面显示不同", "不只是页面回显", "系统执行"]),
        contains_any(normalized, ["DVWA", "实验页面", "目标"]),
    ]

    t3_checks = [
        contains_any(task3, ["Low"]) and contains_any(task3, ["Medium"]),
        contains_any(task3, ["黑名单", "禁止几个符号", "符号限制"]),
        contains_any(task3, ["执行结构", "命令结构", "参数分离"]),
        contains_any(task3, ["SQL 注入", "SQL注入", "共同点", "结构"]),
    ]
    t3_global = [
        contains_any(normalized, ["Low"]) and contains_any(normalized, ["Medium"]),
        contains_any(normalized, ["黑名单", "禁止几个符号", "符号限制"]),
        contains_any(normalized, ["执行结构", "命令结构", "参数分离"]),
        contains_any(normalized, ["SQL 注入", "SQL注入", "共同点", "结构"]),
    ]

    t4_checks = [
        contains_any(task4, ["参数分离", "安全接口", "替代接口"]),
        contains_any(task4, ["最小权限", "最小化暴露"]),
        contains_any(task4, ["结果处理", "边界限制", "后果范围"]),
        count_hits(task4, ["防护", "建议", "控制"]) >= 4 or count_numbered_items(task4) >= 4,
    ]
    t4_global = [
        contains_any(normalized, ["参数分离", "安全接口", "替代接口"]),
        contains_any(normalized, ["最小权限", "最小化暴露"]),
        contains_any(normalized, ["结果处理", "边界限制", "后果范围"]),
        count_hits(normalized, ["防护", "建议", "控制"]) >= 4 or count_numbered_items(normalized) >= 4,
    ]

    task_scores = {
        "task1": max_ratio_score(t1_checks, t1_global, profile.tasks[0].score),
        "task2": max_ratio_score(t2_checks, t2_global, profile.tasks[1].score),
        "task3": max_ratio_score(t3_checks, t3_global, profile.tasks[2].score),
        "task4": max_ratio_score(t4_checks, t4_global, profile.tasks[3].score),
    }

    relation_eval = evaluate_relation_section(
        relation_section,
        ["应用层", "系统层", "输入", "命令", "执行", "结果返回", "参数分离", "最小权限"],
        profile.relation_item_score,
    )
    self_eval_present = bool(self_eval_section) and contains_any(self_eval_section, ["任务完成情况", "机制理解", "证据质量", "自评结论"])
    ai_review_trace = detect_ai_review_trace(ai_review_section or normalized)

    screenshot_markers = count_regex(normalized, SCREENSHOT_MARKER_RE)
    evidence_categories = sum(
        [
            1 if task_scores["task1"] >= 2.2 else 0,
            1 if task_scores["task2"] >= 2.8 else 0,
            1 if task_scores["task3"] >= 2.8 else 0,
            1 if task_scores["task4"] >= 2.8 else 0,
        ]
    )
    artifact_categories = sum(
        [
            1 if contains_any(normalized, ["Command Injection", "命令执行", "应用层", "系统层"]) else 0,
            1 if contains_any(normalized, ["Low", "Medium"]) else 0,
            1 if contains_any(normalized, ["SQL 注入", "SQL注入", "共同点", "结构"]) else 0,
            1 if contains_any(normalized, ["参数分离", "安全接口", "最小权限", "后果范围"]) else 0,
            1 if bool(URL_RE.search(normalized)) or bool(IP_RE.search(normalized)) else 0,
        ]
    )
    evidence_quality = clamp(
        page_points(page_count)
        + screenshot_points(screenshot_markers)
        + ratio_score(evidence_categories, 4, 4.5)
        + ratio_score(artifact_categories, 5, 4.5),
        0.0,
        18.0,
    )

    mechanism_checks = [
        contains_any(normalized, ["应用层"]) and contains_any(normalized, ["系统层"]) and contains_any(normalized, ["跨层"]),
        contains_any(normalized, ["输入"]) and contains_any(normalized, ["命令结构", "执行结构", "拼接"]) and contains_any(normalized, ["执行"]),
        contains_any(normalized, ["Low"]) and contains_any(normalized, ["Medium"]) and contains_any(normalized, ["黑名单", "参数分离"]),
        contains_any(normalized, ["SQL 注入", "SQL注入"]) and contains_any(normalized, ["共同点", "结构"]),
        contains_any(normalized, ["参数分离", "安全接口", "最小权限", "后果范围"]),
    ]
    mechanism_score = ratio_score(sum(mechanism_checks), len(mechanism_checks), 20.0)

    boundary_present = bool(boundary_section) and contains_any(boundary_section, ["授权", "允许", "禁止", "脱敏"])
    risk_checks = [
        boundary_present,
        contains_any(task1 + "\n" + task2 + "\n" + normalized, ["跨层风险", "系统层", "后果", "边界"]),
        contains_any(task4 + "\n" + risk_section, ["最小权限", "后果范围", "边界限制", "安全接口"]),
    ]
    risk_score = ratio_score(sum(risk_checks), len(risk_checks), 12.0)

    control_hits = count_hits(normalized, ["参数分离", "安全接口", "最小权限", "替代接口", "边界限制", "后果范围"])
    protection_score = 10.0 if control_hits >= 6 else (8.0 if control_hits >= 5 else (6.0 if control_hits >= 3 else (3.0 if control_hits >= 2 else 0.0)))

    ordered_sections = all(bool(sections.get(f"task{index}")) for index in range(1, len(profile.tasks) + 1))
    structure_points = 5.0 if ordered_sections and relation_eval["status"] == "complete" else (4.0 if ordered_sections else 2.5)
    text_length = len(normalized)
    expression_points = 5.0 if text_length >= 2400 else (4.0 if text_length >= 1700 else (3.0 if text_length >= 1100 else 1.5))
    expression_score = clamp(structure_points + expression_points, 0.0, 10.0)

    public_targets = detect_public_targets(task2 + "\n" + task3)
    professional_review = build_professional_review(
        normalized,
        ai_review_trace,
        checks=[
            {
                "ok": contains_any(normalized, ["命令结构", "执行结构", "输入进入"]) and not contains_any(normalized, ["记住命令样例就是理解"]),
                "item": "把命令执行理解成记命令样例，而不是输入进入执行结构",
                "why": "第六章的关键是理解输入如何跨层进入系统执行环境，而不是掌握某条命令字符串。",
                "fix": "补写：命令执行的根因在于输入进入了命令结构或系统调用链，而不是记住某条命令本身。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(normalized, ["应用层", "系统层", "跨层"]) and contains_any(normalized, ["结果差异", "系统执行"]),
                "item": "未区分页面差异与系统层执行差异",
                "why": "第六章要求学生看清风险已经越过应用层边界进入系统层。",
                "fix": "补写：观察重点不是页面显示不同，而是输入影响了系统层执行结果。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(normalized, ["SQL 注入", "SQL注入"]) and contains_any(normalized, ["共同点", "结构", "执行环境"]),
                "item": "未建立命令执行与 SQL 注入的共性框架",
                "why": "本章应帮助学生把前面章节迁移到“数据进入执行结构”的统一理解上。",
                "fix": "补写：命令执行与 SQL 注入都属于输入进入执行结构的问题，只是危险环境分别落在系统命令和数据库查询。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(task4 + "\n" + normalized, ["参数分离", "安全接口", "最小权限", "后果范围"]),
                "item": "防护层次区分不足",
                "why": "第六章需要把根因控制、替代调用和后果范围控制区分清楚。",
                "fix": "补写：参数分离和安全接口控制根因，最小权限和边界限制控制后果范围。",
                "penalty": 2.0,
            },
            {
                "ok": boundary_present,
                "item": "授权边界与风险表达不足",
                "why": "命令执行实验必须限定在 DVWA 或教师授权环境，否则会直接越过课程边界。",
                "fix": "补写：实验仅限 DVWA、本机或教师授权环境，禁止对真实未授权系统做验证。",
                "penalty": 2.0,
            },
        ],
    )

    extra_hard_gate_reasons = ["命令执行测试目标授权边界不清"] if public_targets else []
    return finalize_specialized_scoring(
        normalized=normalized,
        page_count=page_count,
        profile=profile,
        sections=sections,
        task_scores=task_scores,
        relation_eval=relation_eval,
        self_eval_present=self_eval_present,
        ai_review_trace=ai_review_trace,
        evidence_quality=evidence_quality,
        mechanism_score=mechanism_score,
        risk_score=risk_score,
        protection_score=protection_score,
        expression_score=expression_score,
        professional_review=professional_review,
        similarity_review=similarity_review,
        rubric_used="projects5_v5_generic_chapter06",
        extra_hard_gate_reasons=extra_hard_gate_reasons,
        extra_debug={
            "screenshot_markers": screenshot_markers,
            "relation_signal_hits": relation_eval["signal_hits"],
            "relation_relation_hits": relation_eval["relation_hits"],
            "public_targets": public_targets[:5],
            "domains": extract_domains(normalized)[:8],
        },
    )


def chapter7_scoring(text: str, page_count: int, profile: ChapterProfile, similarity_review: dict[str, Any] | None) -> tuple[int, dict[str, Any]]:
    normalized = normalize_text(text)
    sections = locate_sections_generic(normalized, profile)

    task1 = sections.get("task1", "") or find_context_window(normalized, ["认证", "会话", "访问控制", "登录到敏感操作"])
    task2 = sections.get("task2", "") or find_context_window(normalized, ["Brute Force", "Low", "Medium", "尝试成本"])
    task3 = sections.get("task3", "") or find_context_window(normalized, ["Weak Session IDs", "Session", "随机", "可预测"])
    task4 = sections.get("task4", "") or find_context_window(normalized, ["CSRF", "来源", "意图", "会话可信"])
    task5 = sections.get("task5", "") or find_context_window(normalized, ["访问控制", "最小权限", "加固清单", "权限"])
    relation_section = sections.get("relation", "") or find_context_window(normalized, ["流程图", "关系图", "认证", "会话", "来源", "权限"])
    boundary_section = sections.get("boundary", "") or find_context_window(normalized, ["授权边界", "授权", "允许", "禁止", "脱敏"])
    risk_section = sections.get("risk", "") or find_context_window(normalized, ["风险", "防护", "控制点", "权限"])
    ai_review_section = sections.get("ai_review", "") or find_context_window(normalized, ["AI输出审核", "人工复核", "人工审核痕迹"])
    self_eval_section = sections.get("self_eval", "") or find_context_window(normalized, ["学生自评表", "自评表", "课程反思"])

    t1_checks = [
        contains_any(task1, ["认证"]) and contains_any(task1, ["会话"]) and contains_any(task1, ["访问控制"]),
        contains_any(task1, ["登录到敏感操作", "流程", "场景"]),
        contains_any(task1, ["认证负责", "会话负责", "访问控制负责"]),
        contains_any(task1, ["请求来源", "来源", "权限"]),
    ]
    t1_global = [
        contains_any(normalized, ["认证"]) and contains_any(normalized, ["会话"]) and contains_any(normalized, ["访问控制"]),
        contains_any(normalized, ["登录到敏感操作", "流程", "场景"]),
        contains_any(normalized, ["认证负责", "会话负责", "访问控制负责"]),
        contains_any(normalized, ["请求来源", "来源", "权限"]),
    ]

    t2_checks = [
        contains_any(task2, ["Brute Force", "暴力破解"]),
        contains_any(task2, ["Low"]) and contains_any(task2, ["Medium"]),
        contains_any(task2, ["尝试成本", "速率限制", "锁定", "验证码"]),
        contains_any(task2, ["认证成本问题", "不是只有弱口令", "控制点"]),
    ]
    t2_global = [
        contains_any(normalized, ["Brute Force", "暴力破解"]),
        contains_any(normalized, ["Low"]) and contains_any(normalized, ["Medium"]),
        contains_any(normalized, ["尝试成本", "速率限制", "锁定", "验证码"]),
        contains_any(normalized, ["认证成本问题", "不是只有弱口令", "控制点"]),
    ]

    t3_checks = [
        contains_any(task3, ["Weak Session", "Session IDs", "Session"]),
        contains_any(task3, ["可预测", "随机", "不可预测"]),
        contains_any(task3, ["持续信任", "会话值", "身份信任链"]),
        contains_any(task3, ["认证与会话", "差异", "不同角色"]),
    ]
    t3_global = [
        contains_any(normalized, ["Weak Session", "Session IDs", "Session"]),
        contains_any(normalized, ["可预测", "随机", "不可预测"]),
        contains_any(normalized, ["持续信任", "会话值", "身份信任链"]),
        contains_any(normalized, ["认证与会话", "差异", "不同角色"]),
    ]

    t4_checks = [
        contains_any(task4, ["CSRF"]),
        contains_any(task4, ["不是密码泄露", "不是密码问题"]),
        contains_any(task4, ["会话可信不等于请求可信", "来源", "意图"]),
        contains_any(task4, ["令牌", "来源校验", "Referer", "Token"]),
    ]
    t4_global = [
        contains_any(normalized, ["CSRF"]),
        contains_any(normalized, ["不是密码泄露", "不是密码问题"]),
        contains_any(normalized, ["会话可信不等于请求可信", "来源", "意图"]),
        contains_any(normalized, ["令牌", "来源校验", "Referer", "Token"]),
    ]

    t5_checks = [
        contains_any(task5, ["访问控制", "权限"]),
        contains_any(task5, ["不能由是否登录替代", "已登录不能替代", "不是登录就够了"]),
        count_hits(task5, ["加固", "建议", "控制"]) >= 4 or count_numbered_items(task5) >= 4,
        contains_any(task5, ["最小权限", "身份", "状态", "来源", "权限"]),
    ]
    t5_global = [
        contains_any(normalized, ["访问控制", "权限"]),
        contains_any(normalized, ["不能由是否登录替代", "已登录不能替代", "不是登录就够了"]),
        count_hits(normalized, ["加固", "建议", "控制"]) >= 4 or count_numbered_items(normalized) >= 4,
        contains_any(normalized, ["最小权限", "身份", "状态", "来源", "权限"]),
    ]

    task_scores = {
        "task1": max_ratio_score(t1_checks, t1_global, profile.tasks[0].score),
        "task2": max_ratio_score(t2_checks, t2_global, profile.tasks[1].score),
        "task3": max_ratio_score(t3_checks, t3_global, profile.tasks[2].score),
        "task4": max_ratio_score(t4_checks, t4_global, profile.tasks[3].score),
        "task5": max_ratio_score(t5_checks, t5_global, profile.tasks[4].score),
    }

    relation_eval = evaluate_relation_section(
        relation_section,
        ["认证", "会话", "来源", "权限", "访问控制", "Brute Force", "CSRF", "最小权限", "流程"],
        profile.relation_item_score,
    )
    self_eval_present = bool(self_eval_section) and contains_any(self_eval_section, ["任务完成情况", "机制理解", "证据质量", "自评结论", "反思"])
    ai_review_trace = detect_ai_review_trace(ai_review_section or normalized)

    screenshot_markers = count_regex(normalized, SCREENSHOT_MARKER_RE)
    evidence_categories = sum(
        [
            1 if task_scores["task1"] >= 2.2 else 0,
            1 if task_scores["task2"] >= 3.0 else 0,
            1 if task_scores["task3"] >= 2.2 else 0,
            1 if task_scores["task4"] >= 2.2 else 0,
            1 if task_scores["task5"] >= 1.5 else 0,
        ]
    )
    artifact_categories = sum(
        [
            1 if contains_any(normalized, ["认证", "会话", "访问控制"]) else 0,
            1 if contains_any(normalized, ["Brute Force", "暴力破解", "Low", "Medium"]) else 0,
            1 if contains_any(normalized, ["Weak Session", "Session IDs", "随机", "可预测"]) else 0,
            1 if contains_any(normalized, ["CSRF", "来源", "意图", "Token"]) else 0,
            1 if contains_any(normalized, ["最小权限", "速率限制", "锁定", "会话保护"]) else 0,
        ]
    )
    evidence_quality = clamp(
        page_points(page_count)
        + screenshot_points(screenshot_markers)
        + ratio_score(evidence_categories, 5, 4.5)
        + ratio_score(artifact_categories, 5, 4.5),
        0.0,
        18.0,
    )

    mechanism_checks = [
        contains_any(normalized, ["认证"]) and contains_any(normalized, ["会话"]) and contains_any(normalized, ["访问控制"]),
        contains_any(normalized, ["尝试成本", "速率限制", "锁定", "验证码"]),
        contains_any(normalized, ["随机", "不可预测", "会话值", "持续信任"]),
        contains_any(normalized, ["CSRF"]) and contains_any(normalized, ["来源", "意图", "会话可信不等于请求可信"]),
        contains_any(normalized, ["已登录"]) and contains_any(normalized, ["不能替代访问控制", "最小权限", "权限"]),
    ]
    mechanism_score = ratio_score(sum(mechanism_checks), len(mechanism_checks), 20.0)

    boundary_present = bool(boundary_section) and contains_any(boundary_section, ["授权", "允许", "禁止", "脱敏"])
    risk_checks = [
        boundary_present,
        contains_any(task2 + "\n" + task3 + "\n" + task4 + "\n" + normalized, ["身份信任链", "尝试成本", "来源", "意图", "持续信任"]),
        contains_any(task5 + "\n" + risk_section, ["访问控制", "最小权限", "权限", "放行边界"]),
    ]
    risk_score = ratio_score(sum(risk_checks), len(risk_checks), 12.0)

    control_hits = count_hits(normalized, ["强密码", "速率限制", "锁定", "验证码", "随机", "令牌", "来源校验", "最小权限", "会话保护"])
    protection_score = 10.0 if control_hits >= 7 else (8.0 if control_hits >= 5 else (6.0 if control_hits >= 4 else (3.0 if control_hits >= 2 else 0.0)))

    ordered_sections = all(bool(sections.get(f"task{index}")) for index in range(1, len(profile.tasks) + 1))
    structure_points = 5.0 if ordered_sections and relation_eval["status"] == "complete" else (4.0 if ordered_sections else 2.5)
    text_length = len(normalized)
    expression_points = 5.0 if text_length >= 2800 else (4.0 if text_length >= 2000 else (3.0 if text_length >= 1300 else 1.5))
    expression_score = clamp(structure_points + expression_points, 0.0, 10.0)

    public_targets = detect_public_targets(task2 + "\n" + task3 + "\n" + task4)
    professional_review = build_professional_review(
        normalized,
        ai_review_trace,
        checks=[
            {
                "ok": contains_any(normalized, ["认证负责", "会话负责", "访问控制负责"]) or (
                    contains_any(normalized, ["认证"]) and contains_any(normalized, ["会话"]) and contains_any(normalized, ["访问控制"]) and contains_any(normalized, ["不同"])
                ),
                "item": "未区分认证、会话与访问控制的职责",
                "why": "第七章要建立身份安全链，而不是把多个概念混成“登录安全”。",
                "fix": "补写：认证回答你是谁，会话回答系统如何继续信任你，访问控制回答你能做什么。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(task2 + "\n" + normalized, ["尝试成本", "速率限制", "锁定"]) and not contains_any(normalized, ["暴力破解只是弱口令问题"]),
                "item": "把暴力破解简化成弱口令问题",
                "why": "第七章要求学生看到认证成本设计，而不是把所有责任都推给用户密码。",
                "fix": "改写为：弱口令会放大风险，但系统允许低成本反复尝试才是暴力破解成立的重要原因。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(task3 + "\n" + normalized, ["随机", "不可预测", "持续信任", "会话值"]),
                "item": "未说明会话随机性与持续信任的关系",
                "why": "本章要把会话值看成持续信任的凭据，而不是普通技术字段。",
                "fix": "补写：会话值若可预测，会直接破坏系统后续继续信任用户的依据。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(task4 + "\n" + normalized, ["不是密码泄露", "来源", "意图", "会话可信不等于请求可信"]),
                "item": "未说明 CSRF 的根因是来源与意图失配",
                "why": "第七章需要把 CSRF 放回“请求来源可信度”这一主线，而不是当成密码问题。",
                "fix": "补写：CSRF 的关键在于系统错误信任了借用用户会话环境发起的请求，而不是攻击者知道了密码。",
                "penalty": 2.0,
            },
            {
                "ok": boundary_present and contains_any(task5 + "\n" + normalized, ["不能由是否登录替代", "最小权限", "访问控制"]),
                "item": "访问控制边界说明不足",
                "why": "第七章最终要收束到“已登录不等于有权执行所有操作”的边界意识。",
                "fix": "补写：访问控制必须独立判断是否允许操作，不能仅以是否登录作为放行依据，并明确实验环境授权边界。",
                "penalty": 2.0,
            },
        ],
    )

    extra_hard_gate_reasons = ["认证/会话相关测试目标授权边界不清"] if public_targets else []
    return finalize_specialized_scoring(
        normalized=normalized,
        page_count=page_count,
        profile=profile,
        sections=sections,
        task_scores=task_scores,
        relation_eval=relation_eval,
        self_eval_present=self_eval_present,
        ai_review_trace=ai_review_trace,
        evidence_quality=evidence_quality,
        mechanism_score=mechanism_score,
        risk_score=risk_score,
        protection_score=protection_score,
        expression_score=expression_score,
        professional_review=professional_review,
        similarity_review=similarity_review,
        rubric_used="projects5_v5_generic_chapter07",
        extra_hard_gate_reasons=extra_hard_gate_reasons,
        extra_debug={
            "screenshot_markers": screenshot_markers,
            "relation_signal_hits": relation_eval["signal_hits"],
            "relation_relation_hits": relation_eval["relation_hits"],
            "public_targets": public_targets[:5],
            "domains": extract_domains(normalized)[:8],
        },
    )


def chapter8_scoring(text: str, page_count: int, profile: ChapterProfile, similarity_review: dict[str, Any] | None) -> tuple[int, dict[str, Any]]:
    normalized = normalize_text(text)
    sections = locate_sections_generic(normalized, profile)

    task1 = sections.get("task1", "") or find_context_window(normalized, ["统一框架", "课程主线", "输入", "处理", "输出", "执行", "身份", "权限"])
    task2 = sections.get("task2", "") or find_context_window(normalized, ["横向比较", "比较表", "SQL", "XSS", "文件上传", "命令执行", "会话"])
    task3 = sections.get("task3", "") or find_context_window(normalized, ["项目计划", "计划单", "主题", "证据来源", "结构安排"])
    task4 = sections.get("task4", "") or find_context_window(normalized, ["项目报告", "展示材料", "展示文稿", "概念", "机制", "证据", "防护"])
    relation_section = sections.get("relation", "") or find_context_window(normalized, ["统一框架图", "结构图", "关系图", "输入", "处理", "输出", "执行", "身份", "权限"])
    boundary_section = sections.get("boundary", "") or find_context_window(normalized, ["授权边界", "引用案例", "授权", "禁止", "脱敏"])
    risk_section = sections.get("risk", "") or find_context_window(normalized, ["互评", "反思", "反馈", "风险", "防护"])
    ai_review_section = sections.get("ai_review", "") or find_context_window(normalized, ["AI输出审核", "人工复核", "人工审核痕迹"])
    self_eval_section = sections.get("self_eval", "") or find_context_window(normalized, ["学生自评表", "自评表", "课程反思", "个人反思"])

    t1_checks = [
        contains_any(task1, ["输入", "处理", "输出", "执行", "身份", "权限"]),
        count_hits(task1, ["第一章", "第二章", "第三章", "第四章", "第五章", "第六章", "第七章", "SQL", "XSS", "文件上传", "命令执行", "认证"]) >= 2,
        contains_any(task1, ["共性", "主线", "统一框架"]),
        contains_any(task1, ["系统化", "迁移", "抽象框架"]),
    ]
    t1_global = [
        contains_any(normalized, ["输入", "处理", "输出", "执行", "身份", "权限"]),
        count_hits(normalized, ["SQL", "XSS", "文件上传", "命令执行", "认证", "会话"]) >= 2,
        contains_any(normalized, ["共性", "主线", "统一框架"]),
        contains_any(normalized, ["系统化", "迁移", "抽象框架"]),
    ]

    t2_checks = [
        count_hits(task2, ["SQL", "XSS", "文件上传", "命令执行", "会话", "访问控制"]) >= 3,
        contains_any(task2, ["比较维度", "输入位置", "危险环境", "根因", "防护重点"]),
        contains_any(task2, ["共性"]) and contains_any(task2, ["差异"]),
        contains_any(task2, ["防护重点", "横向比较", "比较表"]),
    ]
    t2_global = [
        count_hits(normalized, ["SQL", "XSS", "文件上传", "命令执行", "会话", "访问控制"]) >= 3,
        contains_any(normalized, ["比较维度", "输入位置", "危险环境", "根因", "防护重点"]),
        contains_any(normalized, ["共性"]) and contains_any(normalized, ["差异"]),
        contains_any(normalized, ["防护重点", "横向比较", "比较表"]),
    ]

    t3_checks = [
        contains_any(task3, ["项目主题", "主题"]),
        contains_any(task3, ["核心判断", "论点", "要证明什么"]),
        contains_any(task3, ["证据来源", "证据"]),
        contains_any(task3, ["结构安排", "分工", "交付物"]),
    ]
    t3_global = [
        contains_any(normalized, ["项目主题", "主题"]),
        contains_any(normalized, ["核心判断", "论点", "要证明什么"]),
        contains_any(normalized, ["证据来源", "证据"]),
        contains_any(normalized, ["结构安排", "分工", "交付物"]),
    ]

    t4_checks = [
        contains_any(task4, ["概念"]) and contains_any(task4, ["机制"]) and contains_any(task4, ["证据"]) and contains_any(task4, ["防护"]),
        contains_any(task4, ["主论点", "围绕主论点", "结构清楚", "展示材料"]),
        contains_any(task4, ["图表", "比较图", "结构化", "展示"]),
        contains_any(task4, ["可复核", "交付物", "项目报告", "文稿"]),
    ]
    t4_global = [
        contains_any(normalized, ["概念"]) and contains_any(normalized, ["机制"]) and contains_any(normalized, ["证据"]) and contains_any(normalized, ["防护"]),
        contains_any(normalized, ["主论点", "围绕主论点", "结构清楚", "展示材料"]),
        contains_any(normalized, ["图表", "比较图", "结构化", "展示"]),
        contains_any(normalized, ["可复核", "交付物", "项目报告", "文稿"]),
    ]

    task_scores = {
        "task1": max_ratio_score(t1_checks, t1_global, profile.tasks[0].score),
        "task2": max_ratio_score(t2_checks, t2_global, profile.tasks[1].score),
        "task3": max_ratio_score(t3_checks, t3_global, profile.tasks[2].score),
        "task4": max_ratio_score(t4_checks, t4_global, profile.tasks[3].score),
    }

    relation_eval = evaluate_relation_section(
        relation_section,
        ["输入", "处理", "输出", "执行", "身份", "权限", "比较", "项目", "证据", "结构"],
        profile.relation_item_score,
    )
    self_eval_present = bool(self_eval_section) and contains_any(self_eval_section, ["自评", "反思", "互评", "课程反思"])
    ai_review_trace = detect_ai_review_trace(ai_review_section or normalized)

    screenshot_markers = count_regex(normalized, SCREENSHOT_MARKER_RE)
    evidence_categories = sum(
        [
            1 if task_scores["task1"] >= 2.6 else 0,
            1 if task_scores["task2"] >= 2.6 else 0,
            1 if task_scores["task3"] >= 2.2 else 0,
            1 if task_scores["task4"] >= 3.8 else 0,
        ]
    )
    artifact_categories = sum(
        [
            1 if contains_any(normalized, ["输入", "处理", "输出", "执行", "身份", "权限"]) else 0,
            1 if count_hits(normalized, ["SQL", "XSS", "文件上传", "命令执行", "会话", "访问控制"]) >= 3 else 0,
            1 if contains_any(normalized, ["比较表", "横向比较", "比较维度"]) else 0,
            1 if contains_any(normalized, ["项目主题", "论点", "证据来源", "交付物"]) else 0,
            1 if contains_any(normalized, ["互评", "反思", "反馈"]) else 0,
        ]
    )
    evidence_quality = clamp(
        page_points(page_count)
        + screenshot_points(screenshot_markers)
        + ratio_score(evidence_categories, 4, 5.0)
        + ratio_score(artifact_categories, 5, 4.0),
        0.0,
        18.0,
    )

    mechanism_checks = [
        contains_any(normalized, ["统一框架", "输入", "处理", "输出", "执行", "身份", "权限"]),
        count_hits(normalized, ["SQL", "XSS", "文件上传", "命令执行", "会话"]) >= 3 and contains_any(normalized, ["共性", "差异"]),
        contains_any(normalized, ["目标", "证据", "结构", "交付物"]),
        contains_any(normalized, ["概念", "机制", "证据", "防护"]),
        contains_any(normalized, ["互评", "反思", "反馈", "思维变化"]),
    ]
    mechanism_score = ratio_score(sum(mechanism_checks), len(mechanism_checks), 20.0)

    boundary_present = bool(boundary_section) and contains_any(boundary_section, ["授权", "案例", "引用", "脱敏"])
    risk_checks = [
        boundary_present,
        contains_any(task3 + "\n" + task4 + "\n" + normalized, ["目标", "证据", "结构", "交付物"]),
        contains_any(risk_section + "\n" + normalized, ["互评", "反思", "反馈", "课程复盘"]),
    ]
    risk_score = ratio_score(sum(risk_checks), len(risk_checks), 12.0)

    control_hits = count_hits(normalized, ["防护", "加固", "参数化查询", "输出编码", "隔离存储", "最小权限", "令牌", "速率限制"])
    protection_score = 10.0 if control_hits >= 7 else (8.0 if control_hits >= 5 else (6.0 if control_hits >= 3 else (3.0 if control_hits >= 2 else 0.0)))

    ordered_sections = all(bool(sections.get(f"task{index}")) for index in range(1, len(profile.tasks) + 1))
    structure_points = 5.0 if ordered_sections and relation_eval["status"] == "complete" else (4.0 if ordered_sections else 2.5)
    text_length = len(normalized)
    expression_points = 5.0 if text_length >= 3000 else (4.0 if text_length >= 2200 else (3.0 if text_length >= 1500 else 1.5))
    expression_score = clamp(structure_points + expression_points, 0.0, 10.0)

    public_targets = detect_public_targets(normalized) if contains_any(normalized, ["测试", "验证", "扫描", "抓包", "攻击"]) else []
    professional_review = build_professional_review(
        normalized,
        ai_review_trace,
        checks=[
            {
                "ok": contains_any(normalized, ["统一框架", "输入", "处理", "输出", "执行", "身份", "权限"]),
                "item": "未真正使用统一框架组织前面章节",
                "why": "第八章的核心不是重新罗列章节，而是把前面内容组织成统一分析模型。",
                "fix": "补写：至少用“输入—处理—输出—执行—身份—权限”框架把多章内容挂接起来。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(task2 + "\n" + normalized, ["比较维度", "共性", "差异", "防护重点"]),
                "item": "漏洞比较缺少固定维度",
                "why": "第八章要训练横向比较，而不是平铺罗列章节结论。",
                "fix": "补写：围绕输入位置、危险环境、核心根因、防护重点等固定维度做比较。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(task3 + "\n" + task4 + "\n" + normalized, ["目标", "证据", "结构", "交付物"]),
                "item": "项目论证链不完整",
                "why": "综合项目不是堆内容，而是目标、证据、结构和交付物的对齐。",
                "fix": "补写：项目要证明什么、证据从哪里来、如何组织结构、最终交付什么。",
                "penalty": 2.0,
            },
            {
                "ok": contains_any(task4 + "\n" + normalized, ["概念", "机制", "证据", "防护"]),
                "item": "展示材料结构不完整",
                "why": "第八章展示应把概念、机制、证据、防护串成论证，而不是只做现象罗列。",
                "fix": "补写：每个案例或比较项都要回答是什么、为什么、证据在哪、如何控制。",
                "penalty": 2.0,
            },
            {
                "ok": boundary_present and contains_any(normalized, ["互评", "反思", "反馈", "AI输出审核"]),
                "item": "案例边界或复盘表达不足",
                "why": "综合复盘章节要求学生对案例来源、授权边界、互评反馈和 AI 使用审核做负责任表达。",
                "fix": "补写：案例和截图来源、授权边界、互评依据、个人反思，以及 AI 输出审核与删改说明。",
                "penalty": 2.0,
            },
        ],
    )

    extra_hard_gate_reasons = ["综合项目引用或测试目标授权边界不清"] if public_targets else []
    return finalize_specialized_scoring(
        normalized=normalized,
        page_count=page_count,
        profile=profile,
        sections=sections,
        task_scores=task_scores,
        relation_eval=relation_eval,
        self_eval_present=self_eval_present,
        ai_review_trace=ai_review_trace,
        evidence_quality=evidence_quality,
        mechanism_score=mechanism_score,
        risk_score=risk_score,
        protection_score=protection_score,
        expression_score=expression_score,
        professional_review=professional_review,
        similarity_review=similarity_review,
        rubric_used="projects5_v5_generic_chapter08",
        extra_hard_gate_reasons=extra_hard_gate_reasons,
        extra_debug={
            "screenshot_markers": screenshot_markers,
            "relation_signal_hits": relation_eval["signal_hits"],
            "relation_relation_hits": relation_eval["relation_hits"],
            "public_targets": public_targets[:5],
            "domains": extract_domains(normalized)[:8],
        },
    )


def generic_fallback_scoring(text: str, page_count: int, profile: ChapterProfile, similarity_review: dict[str, Any] | None) -> tuple[int, dict[str, Any]]:
    normalized = normalize_text(text)
    sections = locate_sections_generic(normalized, profile)
    ai_review_trace = detect_ai_review_trace(sections.get("ai_review", "") or normalized)
    self_eval_present = bool(sections.get("self_eval")) and contains_any(sections.get("self_eval", ""), ["自评", "任务完成"])
    relation_eval = evaluate_relation_section(
        sections.get("relation", ""),
        [profile.relation_item_name] + profile.capability_goals[:4],
        profile.relation_item_score,
    )

    task_scores: dict[str, float] = {}
    for index, task in enumerate(profile.tasks, start=1):
        section_text = sections.get(f"task{index}", "")
        requirement_hits = 0
        for requirement in task.semantic_requirements:
            keywords = [part for part in re.split(r"[、，,；/（）() ]+", requirement) if len(part.strip()) >= 2]
            if contains_any(section_text or normalized, keywords):
                requirement_hits += 1
        task_scores[f"task{index}"] = ratio_score(requirement_hits, max(1, len(task.semantic_requirements)), task.score)

    task_completion = clamp(sum(task_scores.values()) + relation_eval["score"] + (profile.self_eval_score if self_eval_present else 0.0), 0.0, 20.0)
    screenshot_markers = count_regex(normalized, SCREENSHOT_MARKER_RE)
    evidence_quality = clamp(page_points(page_count) + screenshot_points(screenshot_markers) + ratio_score(sum(1 for value in task_scores.values() if value > 0), len(task_scores), 9.0), 0.0, 18.0)
    mechanism_score = ratio_score(sum(1 for goal in profile.capability_goals if contains_any(normalized, [goal[:4], goal[:6]])), max(1, len(profile.capability_goals)), 20.0)
    risk_score = 8.0 if contains_any(normalized, ["授权", "脱敏", "风险", "防护"]) else 4.0
    protection_score = 6.0 if contains_any(normalized, ["防护", "建议", "控制"]) else 2.0
    text_length = len(normalized)
    expression_score = 10.0 if text_length >= 1800 else (7.0 if text_length >= 1000 else 4.0)
    professional_review = build_professional_review(normalized, ai_review_trace, [])
    professional_score = professional_review["score"]

    content_issues: list[str] = []
    for key, value in task_scores.items():
        max_score = next(task.score for index, task in enumerate(profile.tasks, start=1) if f"task{index}" == key)
        if value < 0.6 * max_score:
            content_issues.append(f"{profile.tasks[int(key[-1]) - 1].name}内容不全")
    if relation_eval["status"] == "missing":
        content_issues.append(f"未提交{profile.relation_item_name}")
    if not ai_review_trace["present"]:
        content_issues.append("缺少AI输出审核与人工复核记录")
    hard_gate_reasons: list[str] = []
    if SESSION_LEAK_RE.search(normalized) and not MASK_HINT_RE.search(normalized):
        hard_gate_reasons.append("敏感信息疑似未脱敏")
    if similarity_review and similarity_review.get("is_highly_similar"):
        hard_gate_reasons.append("与同班作业高度雷同")
    final_score = int(round(clamp(task_completion + evidence_quality + mechanism_score + risk_score + protection_score + expression_score + professional_score, 0.0, 100.0)))
    if hard_gate_reasons:
        final_score = min(final_score, 59)
    debug = {
        "rubric_used": "projects5_v5_generic_fallback",
        "task_completion": round(task_completion, 2),
        "evidence_quality": round(evidence_quality, 2),
        "mechanism_score": round(mechanism_score, 2),
        "risk_score": round(risk_score, 2),
        "protection_score": round(protection_score, 2),
        "expression_score": round(expression_score, 2),
        "professional_score": round(professional_score, 2),
        "final_score": final_score,
        "relation_status": relation_eval["status"],
        "self_eval_present": self_eval_present,
        "ai_review_present": ai_review_trace["present"],
        "task_scores": {key: round(value, 2) for key, value in task_scores.items()},
        "task_max_scores": {f"task{index}": task.score for index, task in enumerate(profile.tasks, start=1)},
        "task_name_map": {f"task{index}": task.name for index, task in enumerate(profile.tasks, start=1)},
        "task_requirement_map": {f"task{index}": task.semantic_requirements for index, task in enumerate(profile.tasks, start=1)},
        "content_issues": content_issues,
        "hard_gate_reasons": hard_gate_reasons,
        "ai_like_paragraphs": professional_review["ai_like_paragraphs"],
        "unchecked_traces": professional_review["unchecked_traces"],
        "professional_errors": professional_review["professional_errors"],
        "absolute_hits": professional_review["absolute_hits"],
        "similarity_review": similarity_review or {
            "is_suspected_similar": False,
            "is_highly_similar": False,
            "max_full_ratio": 0.0,
            "max_longest_block_ratio": 0.0,
            "matched_peers": [],
        },
    }
    return final_score, debug


def score_text(text: str, page_count: int, profile: ChapterProfile, similarity_review: dict[str, Any] | None = None) -> tuple[int, dict[str, Any]]:
    if "第一章" in profile.chapter_name:
        return chapter1_scoring(text, page_count, profile, similarity_review)
    if "第二章" in profile.chapter_name:
        return chapter2_scoring(text, page_count, profile, similarity_review)
    if "第三章" in profile.chapter_name or "SQL注入" in profile.chapter_name:
        return chapter3_scoring(text, page_count, profile, similarity_review)
    if "第四章" in profile.chapter_name or "XSS" in profile.chapter_name:
        return chapter4_scoring(text, page_count, profile, similarity_review)
    if "第五章" in profile.chapter_name or "文件上传" in profile.chapter_name:
        return chapter5_scoring(text, page_count, profile, similarity_review)
    if "第六章" in profile.chapter_name or "命令执行" in profile.chapter_name:
        return chapter6_scoring(text, page_count, profile, similarity_review)
    if "第七章" in profile.chapter_name or "身份认证" in profile.chapter_name or "访问控制" in profile.chapter_name:
        return chapter7_scoring(text, page_count, profile, similarity_review)
    if "第八章" in profile.chapter_name or "综合复盘" in profile.chapter_name or "项目展示" in profile.chapter_name:
        return chapter8_scoring(text, page_count, profile, similarity_review)
    return generic_fallback_scoring(text, page_count, profile, similarity_review)


def score_pdf_file(pdf_path: Path, profile: ChapterProfile, extractor: OCRExtractor) -> SubmissionOutcome:
    text, page_count = extractor.extract_pdf_text(pdf_path)
    score, debug = score_text(text, page_count, profile)
    student_id, student_name = extract_submission_identity(pdf_path)
    label = label_for_score(score, debug.get("hard_gate_reasons", []))
    debug["label"] = label
    return SubmissionOutcome(
        file_path=str(pdf_path),
        student_id=student_id,
        student_name=student_name,
        page_count=page_count,
        submitted=True,
        score=score,
        label=label,
        assessment=build_assessment(score, debug),
        good_points=build_good_points(debug),
        suggestion=build_suggestion(profile, debug, submitted=True),
        debug=debug,
    )


def build_missing_outcome(profile: ChapterProfile, student_id: str, student_name: str) -> SubmissionOutcome:
    debug = {
        "rubric_used": "projects5_v5_missing_submission",
        "task_completion": 0.0,
        "evidence_quality": 0.0,
        "mechanism_score": 0.0,
        "risk_score": 0.0,
        "protection_score": 0.0,
        "expression_score": 0.0,
        "professional_score": 0.0,
        "base_score": 0.0,
        "final_score": 0,
        "relation_status": "missing",
        "self_eval_present": False,
        "ai_review_present": False,
        "task_scores": {f"task{index}": 0.0 for index, _task in enumerate(profile.tasks, start=1)},
        "task_max_scores": {f"task{index}": task.score for index, task in enumerate(profile.tasks, start=1)},
        "task_name_map": {f"task{index}": task.name for index, task in enumerate(profile.tasks, start=1)},
        "task_requirement_map": {f"task{index}": task.semantic_requirements for index, task in enumerate(profile.tasks, start=1)},
        "content_issues": ["未提交本章作业"],
        "hard_gate_reasons": ["未提交本章作业"],
        "label": "需补交",
    }
    return SubmissionOutcome(
        file_path=None,
        student_id=student_id,
        student_name=student_name,
        page_count=0,
        submitted=False,
        score=0,
        label="需补交",
        assessment="0分，需补交；未提交本章作业。",
        good_points="未提交，本项暂无可评优点。",
        suggestion=build_suggestion(profile, debug, submitted=False),
        debug=debug,
    )


def load_roster(excel_path: Path) -> list[tuple[str, str]]:
    wb = load_workbook(excel_path)
    ws = wb.active
    roster: list[tuple[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        student_id = str(row[0]).strip() if row and row[0] is not None else ""
        student_name = str(row[1]).strip() if row and len(row) > 1 and row[1] is not None else ""
        if student_id and student_name:
            roster.append((student_id, student_name))
    return roster


def score_pdf_directory(
    pdf_dir: Path,
    profile: ChapterProfile,
    extractor: OCRExtractor,
    roster: list[tuple[str, str]] | None = None,
) -> dict[str, Any]:
    pdf_paths = [path for path in sorted(pdf_dir.glob("*.pdf")) if "评估" not in path.stem]
    extracted: dict[str, tuple[Path, str, int]] = {}
    for pdf_path in pdf_paths:
        _student_id, student_name = extract_submission_identity(pdf_path)
        text, page_count = extractor.extract_pdf_text(pdf_path)
        key = student_name or pdf_path.stem
        extracted[key] = (pdf_path, text, page_count)

    similarity_reviews = detect_similarity_reviews(profile, {name: payload[1] for name, payload in extracted.items()})
    results: list[SubmissionOutcome] = []
    for name, (pdf_path, text, page_count) in extracted.items():
        score, debug = score_text(text, page_count, profile, similarity_review=similarity_reviews.get(name))
        student_id, student_name = extract_submission_identity(pdf_path)
        label = label_for_score(score, debug.get("hard_gate_reasons", []))
        debug["label"] = label
        results.append(
            SubmissionOutcome(
                file_path=str(pdf_path),
                student_id=student_id,
                student_name=student_name,
                page_count=page_count,
                submitted=True,
                score=score,
                label=label,
                assessment=build_assessment(score, debug),
                good_points=build_good_points(debug),
                suggestion=build_suggestion(profile, debug, submitted=True),
                debug=debug,
            )
        )

    if roster:
        result_map = {(item.student_id, item.student_name): item for item in results if item.student_id and item.student_name}
        name_map = {item.student_name: item for item in results if item.student_name}
        ordered_results: list[SubmissionOutcome] = []
        used_names: set[str] = set()
        for student_id, student_name in roster:
            matched = result_map.get((student_id, student_name)) or name_map.get(student_name)
            if matched:
                ordered_results.append(matched)
                used_names.add(student_name)
            else:
                ordered_results.append(build_missing_outcome(profile, student_id, student_name))
        extras = [item for item in results if item.student_name not in used_names]
        results = ordered_results + extras

    summary = {
        "profile": summarize_profile(profile),
        "file_count": len(results),
        "submitted_count": sum(1 for item in results if item.submitted),
        "missing_count": sum(1 for item in results if not item.submitted),
        "avg_score": round(sum(item.score for item in results) / len(results), 2) if results else 0.0,
        "max_score": max((item.score for item in results), default=0),
        "min_score": min((item.score for item in results), default=0),
        "high_similarity_count": sum(1 for item in results if item.debug.get("similarity_review", {}).get("is_highly_similar")),
    }
    return {
        "summary": summary,
        "results": [
            {
                "file_path": item.file_path,
                "student_id": item.student_id,
                "student_name": item.student_name,
                "page_count": item.page_count,
                "submitted": item.submitted,
                "score": item.score,
                "label": item.label,
                "assessment": item.assessment,
                "good_points": item.good_points,
                "suggestion": item.suggestion,
                "debug": item.debug,
            }
            for item in results
        ],
    }


def update_excel(excel_path: Path, outcomes: list[SubmissionOutcome], profile: ChapterProfile) -> None:
    wb = load_workbook(excel_path)
    ws = wb.active
    score_header = f"{profile.chapter_name} v5.0分数"
    label_header = f"{profile.chapter_name} v5.0等级"
    good_header = "做得好的地方"
    suggestion_header = "需要完善的地方"
    headers = {3: score_header, 4: label_header, 5: good_header, 6: suggestion_header}
    for col, title in headers.items():
        ws.cell(row=1, column=col).value = title

    outcome_map = {(item.student_id, item.student_name): item for item in outcomes}
    for row_index in range(2, ws.max_row + 1):
        student_id = str(ws.cell(row=row_index, column=1).value).strip() if ws.cell(row=row_index, column=1).value is not None else ""
        student_name = str(ws.cell(row=row_index, column=2).value).strip() if ws.cell(row=row_index, column=2).value is not None else ""
        outcome = outcome_map.get((student_id, student_name))
        ws.cell(row=row_index, column=3).value = outcome.score if outcome else 0
        ws.cell(row=row_index, column=4).value = outcome.label if outcome else "需补交"
        ws.cell(row=row_index, column=5).value = outcome.good_points if outcome else "未提交，本项暂无可评优点。"
        ws.cell(row=row_index, column=6).value = outcome.suggestion if outcome else build_suggestion(profile, {"hard_gate_reasons": ["未提交本章作业"]}, submitted=False)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 44
    wb.save(excel_path)


def build_pdf(output_path: Path, profile: ChapterProfile, class_name: str, outcomes: list[SubmissionOutcome]) -> None:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="CN",
            fontName="STSong-Light",
            fontSize=9,
            leading=13,
            textColor=colors.HexColor("#1f2937"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="CNTitle",
            fontName="STSong-Light",
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#111827"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="CNMeta",
            fontName="STSong-Light",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#4b5563"),
        )
    )

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    submitted_count = sum(1 for item in outcomes if item.submitted)
    missing_count = len(outcomes) - submitted_count

    def label_count(label: str) -> int:
        return sum(1 for item in outcomes if item.label == label)

    story = [
        Paragraph(f"{class_name} {profile.chapter_name} 作业评估", styles["CNTitle"]),
        Spacer(1, 4 * mm),
        Paragraph(
            f"评分依据：projects5.0 课程级母规则 + 章节 profile（{profile.chapter_name}）。"
            "本 PDF 不显示具体分数，只展示等级、做得好的地方和需要完善的地方，便于学生复盘与教师反馈。",
            styles["CNMeta"],
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            f"班级人数：{len(outcomes)}；已提交：{submitted_count}；未提交：{missing_count}",
            styles["CNMeta"],
        ),
        Spacer(1, 5 * mm),
    ]

    table_data = [
        [
            Paragraph("学号", styles["CN"]),
            Paragraph("姓名", styles["CN"]),
            Paragraph("等级", styles["CN"]),
            Paragraph("做得好的地方", styles["CN"]),
            Paragraph("需要完善的地方", styles["CN"]),
        ]
    ]
    for item in outcomes:
        table_data.append(
            [
                Paragraph(item.student_id or "-", styles["CN"]),
                Paragraph(item.student_name or "-", styles["CN"]),
                Paragraph(item.label, styles["CN"]),
                Paragraph(item.good_points, styles["CN"]),
                Paragraph(item.suggestion, styles["CN"]),
            ]
        )

    table = LongTable(table_data, colWidths=[16 * mm, 22 * mm, 18 * mm, 58 * mm, 66 * mm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("LEADING", (0, 0), (-1, -1), 12),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#94a3b8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("评分摘要", styles["CNTitle"]))
    story.append(Spacer(1, 3 * mm))
    story.append(
        Paragraph(
            f"优秀：{label_count('优秀')}；良好：{label_count('良好')}；中等：{label_count('中等')}；及格：{label_count('及格')}；待补救：{label_count('待补救')}；需补交：{label_count('需补交')}",
            styles["CNMeta"],
        )
    )
    doc.build(story)


def write_debug_json(debug_path: Path, payload: dict[str, Any]) -> None:
    debug_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def default_output_path(base_dir: Path, profile: ChapterProfile, suffix: str) -> Path:
    chapter_stub = re.sub(r"[^\w\u4e00-\u9fff]+", "_", profile.chapter_name).strip("_")
    if not chapter_stub:
        chapter_stub = "chapter"
    return base_dir / f"{chapter_stub}_v5{suffix}"


def main() -> None:
    parser = argparse.ArgumentParser(description="projects5.0 通用作业评估引擎")
    parser.add_argument("--profile", type=Path, help="章节变量配置 JSON 路径")
    parser.add_argument("--show-schema", action="store_true", help="输出 profile 所需字段示例")
    parser.add_argument("--pdf", type=Path, help="对单个 PDF 作业评分")
    parser.add_argument("--dir", type=Path, help="对目录下所有 PDF 作业评分")
    parser.add_argument("--excel", type=Path, help="班级 roster Excel；若提供，将按学号和姓名回写分数与等级")
    parser.add_argument("--pdf-output", type=Path, help="评估 PDF 输出路径")
    parser.add_argument("--debug-output", type=Path, help="debug JSON 输出路径")
    args = parser.parse_args()

    if args.show_schema:
        schema = {
            "course_name": "Web应用安全与防护",
            "chapter_name": "第二章 Web信息收集与请求分析",
            "chapter_mainline": "从线索到证据再到授权探测",
            "capability_goals": ["区分线索/证据/主动探测", "解释工具为什么得到这类结果"],
            "tasks": [
                {
                    "name": "Google 公开页面线索记录",
                    "score": 2.5,
                    "required": True,
                    "semantic_requirements": ["查询语句", "至少一种搜索语法", "为什么只能算线索"],
                }
            ],
            "relation_item_name": "综合关系图/表",
            "relation_item_score": 3,
            "self_eval_score": 2,
            "redlines": [
                {"name": "未脱敏", "description": "敏感信息未脱敏", "action": "lt_60"},
                {"name": "超出授权边界", "description": "对未授权目标执行测试", "action": "lt_60"},
            ],
            "professional_checks": ["是否区分线索与证据", "是否存在绝对化表述"],
            "default_tags": ["T1", "E1", "M1", "R2", "B1"],
        }
        print(json.dumps(schema, ensure_ascii=False, indent=2))
        return

    if not args.profile:
        raise SystemExit("请通过 --profile 指定章节变量配置 JSON。")

    profile = load_profile(args.profile)
    errors = validate_profile(profile)
    if errors:
        print(
            json.dumps(
                {
                    "summary": summarize_profile(profile),
                    "validation_errors": errors,
                    "profile": asdict(profile),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        raise SystemExit(1)

    if not args.pdf and not args.dir:
        print(
            json.dumps(
                {
                    "summary": summarize_profile(profile),
                    "validation_errors": [],
                    "profile": asdict(profile),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    extractor = OCRExtractor()

    if args.pdf:
        if args.excel or args.pdf_output or args.debug_output:
            raise SystemExit("单 PDF 评分模式暂不支持 Excel/PDF 批量输出，请使用 --dir。")
        outcome = score_pdf_file(args.pdf, profile, extractor)
        print(
            json.dumps(
                {
                    "profile": summarize_profile(profile),
                    "result": {
                        "file_path": outcome.file_path,
                        "student_id": outcome.student_id,
                        "student_name": outcome.student_name,
                        "page_count": outcome.page_count,
                        "score": outcome.score,
                        "label": outcome.label,
                        "assessment": outcome.assessment,
                        "good_points": outcome.good_points,
                        "suggestion": outcome.suggestion,
                        "debug": outcome.debug,
                    },
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    if args.dir:
        roster = load_roster(args.excel) if args.excel else None
        payload = score_pdf_directory(args.dir, profile, extractor, roster=roster)
        outcomes = [
            SubmissionOutcome(
                file_path=item.get("file_path"),
                student_id=item.get("student_id", ""),
                student_name=item.get("student_name", ""),
                page_count=item.get("page_count", 0),
                submitted=item.get("submitted", True),
                score=item.get("score", 0),
                label=item.get("label", ""),
                assessment=item.get("assessment", ""),
                good_points=item.get("good_points", ""),
                suggestion=item.get("suggestion", ""),
                debug=item.get("debug", {}),
            )
            for item in payload["results"]
        ]

        if args.excel:
            update_excel(args.excel, outcomes, profile)
            payload["excel_updated"] = str(args.excel)

        if args.pdf_output:
            pdf_output = args.pdf_output
        elif args.excel:
            pdf_output = default_output_path(args.excel.parent, profile, "_评估.pdf")
        else:
            pdf_output = None

        if pdf_output:
            class_name = args.dir.name
            build_pdf(pdf_output, profile, class_name, outcomes)
            payload["pdf_output"] = str(pdf_output)

        if args.debug_output:
            debug_output = args.debug_output
        elif args.excel:
            debug_output = default_output_path(args.excel.parent, profile, "_debug.json")
        else:
            debug_output = None

        if debug_output:
            write_debug_json(debug_output, payload)
            payload["debug_output"] = str(debug_output)

        print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
